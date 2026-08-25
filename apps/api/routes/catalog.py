"""Workspace catalog API: insumos, APU matrices, and rendimientos.

Edits here change the workspace baseline used by every future (re)compute;
per-project price overrides in costing_overrides.json still apply on top.
A global ``catalog_updated`` event tells open projects to offer a recompute.
"""

import csv
import hashlib
import io
from pathlib import Path
from typing import Annotated, Literal

from fastapi import APIRouter, Depends, Header, HTTPException, Request, Response, UploadFile
from klave_engine.common.config import Settings
from klave_engine.common.errors import ReportGenerationError
from klave_engine.costing.apu import build_all_apus
from klave_engine.costing.catalog import PHASE_ORDER, build_catalog_from_store
from klave_engine.costing.catalog_services import (
    apply_equipment,
    apply_labor,
    import_plantilla,
    labor_state,
)
from klave_engine.costing.catalog_store import CatalogStore, UnitMismatch
from klave_engine.costing.descripciones import long_description
from klave_engine.costing.equipment import EquipmentParameters, compute_costo_horario
from klave_engine.costing.exports import build_cotizacion_workbook
from klave_engine.costing.labor import (
    REGION_PRESETS,
    FsrParameters,
    LaborCategory,
    apply_preset,
    compute_fsr,
    preset_by_key,
)
from klave_engine.costing.matching import Candidate, Match, rank, split_alcance
from klave_engine.costing.models import CostingAssumptions, CostingOverrides
from klave_engine.costing.recompute import load_overrides, recompute_and_persist
from klave_engine.costing.sources.custom import (
    CustomCatalogError,
    parse_concept_workbook,
    source_key_for,
)
from klave_engine.costing.sources.matrices import parse_matrices_workbook
from klave_engine.costing.sources.registry import SOURCES, available_sources, sources_dir
from klave_engine.costing.vigencia import (
    FRESH_MONTHS,
    STALE_MONTHS,
    now_month,
    price_ages,
    roll_forward_factor,
)
from pydantic import BaseModel, Field

from apps.api.auth.common import rate_limit
from apps.api.dependencies import ProjectStore, get_settings, get_store
from apps.api.events import BUS, clean_actor
from apps.api.routes.projects import visible_project_filter
from apps.api.tenancy import store_for_project, store_for_request

router = APIRouter(prefix="/catalog", tags=["catalog"])

MAX_IMPORT_BYTES = 1_000_000


def require_catalog_admin(request: Request) -> None:
    """Structural catálogo changes — imports, deletions, repricing sweeps,
    labor/equipment engines — are the administrator's. Daily estimator work
    (a price from a cotización, an alias, a mapping, a factor) stays open to
    every active member: it is their own taller's catálogo. Open mode (no
    accounts) keeps the local-first freedom."""
    user = getattr(request.state, "user", None)
    if user is not None and user.get("role") != "admin":
        raise HTTPException(
            status_code=403,
            detail={
                "error_type": "admin_required",
                "message": "Esta operación del catálogo requiere un administrador del taller.",
            },
        )


def get_catalog(request: Request, settings: Settings = Depends(get_settings)) -> CatalogStore:
    """The signed-in taller's own catálogo; the shared one only in open mode."""
    return store_for_request(settings, request)


class InsumoUpdate(BaseModel):
    description: str | None = Field(default=None, min_length=1, max_length=200)
    unit: str | None = Field(default=None, min_length=1, max_length=12)
    resource_type: str | None = None
    unit_cost: float | None = Field(default=None, gt=0)
    source: str | None = Field(default=None, max_length=200)
    source_type: Literal["referencia", "cotizacion", "publicacion", "calculado"] | None = None
    region: str | None = Field(default=None, max_length=12)
    vigencia: str | None = Field(default=None, max_length=7)


class ConceptCreate(BaseModel):
    code: str = Field(min_length=2, max_length=40, pattern=r"^[A-Za-z0-9._-]+$")
    description: str = Field(min_length=3, max_length=200)
    unit: str = Field(min_length=1, max_length=12)
    phase: str = Field(min_length=2, max_length=40)
    production_rate_per_day: float = Field(gt=0)
    components: list["ApuComponentInput"] = Field(min_length=1)


class ConceptUpdate(BaseModel):
    description: str | None = Field(default=None, min_length=3, max_length=200)
    unit: str | None = Field(default=None, min_length=1, max_length=12)
    phase: str | None = Field(default=None, min_length=2, max_length=40)
    production_rate_per_day: float | None = Field(default=None, gt=0)
    active: bool | None = None


class InsumoCreate(BaseModel):
    code: str = Field(min_length=2, max_length=40, pattern=r"^[A-Za-z0-9._-]+$")
    description: str = Field(min_length=1, max_length=200)
    unit: str = Field(min_length=1, max_length=12)
    resource_type: str
    unit_cost: float = Field(gt=0)
    source: str = Field(default="", max_length=200)


class ApuComponentInput(BaseModel):
    resource_code: str
    quantity: float = Field(gt=0)


class ApuUpdate(BaseModel):
    components: list[ApuComponentInput] = Field(min_length=1)


class RendimientoUpdate(BaseModel):
    production_rate_per_day: float = Field(gt=0)


def _publish_catalog_updated(
    actor: str | None, action: str, detail: str, catalog: CatalogStore | None = None
) -> None:
    BUS.publish(
        "catalog_updated",
        actor=clean_actor(actor),
        data={"action": action, "detail": detail},
        # The event stays inside the taller whose catálogo changed.
        workspace_id=getattr(catalog, "workspace_id", None),
    )


@router.get("")
def get_catalog_state(catalog: CatalogStore = Depends(get_catalog)) -> dict:
    templates = catalog.load_templates()
    return {
        "insumos": catalog.list_insumos(),
        "concepts": [
            {
                "code": row["code"],
                "description": row["description"],
                "unit": row["unit"],
                "phase": row["phase"],
                "production_rate_per_day": row["production_rate_per_day"],
                # rule_key marks detection-backed concepts; without one the
                # concept is manual and quantities come from adjustments.
                "detection_backed": bool(row.get("rule_key")),
                # A P.U. adopted from a reference row replaces the matrix.
                "price_override": row.get("price_override"),
                "price_source": row.get("price_source"),
                "price_clave": row.get("price_clave"),
                "price_vigencia": row.get("price_vigencia"),
            }
            for row in catalog.load_concepts()
        ],
        "apus": {
            code: [
                {"resource_code": resource_code, "quantity": quantity}
                for resource_code, quantity in components
            ]
            for code, components in templates.items()
        },
        "phase_order": PHASE_ORDER,
    }


@router.post("/insumos", status_code=201)
def create_insumo(
    body: InsumoCreate,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    if body.code in catalog.load_price_book():
        raise HTTPException(
            status_code=409,
            detail={"error_type": "insumo_exists", "code": body.code},
        )
    try:
        row = catalog.upsert_insumo(
            body.code,
            description=body.description,
            unit=body.unit,
            resource_type=body.resource_type,
            unit_cost=body.unit_cost,
            source=body.source,
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=422, detail={"error_type": "invalid_insumo", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(x_actor, "insumo_created", body.code, catalog=catalog)
    return row


@router.put("/insumos/{code}")
def update_insumo(
    code: str,
    body: InsumoUpdate,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    if code not in catalog.load_price_book():
        raise HTTPException(
            status_code=404, detail={"error_type": "insumo_not_found", "code": code}
        )
    try:
        row = catalog.upsert_insumo(
            code,
            description=body.description,
            unit=body.unit,
            resource_type=body.resource_type,
            unit_cost=body.unit_cost,
            source=body.source,
            source_type=body.source_type,
            region=body.region,
            vigencia=body.vigencia,
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=422, detail={"error_type": "invalid_insumo", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(x_actor, "insumo_updated", code, catalog=catalog)
    return row


@router.put("/apus/{concept_code}")
def update_apu(
    concept_code: str,
    body: ApuUpdate,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    known = {row["code"] for row in catalog.load_concepts(include_inactive=True)}
    if concept_code not in known:
        raise HTTPException(
            status_code=404,
            detail={"error_type": "concept_not_found", "code": concept_code},
        )
    try:
        catalog.set_apu_components(
            concept_code,
            [(component.resource_code, component.quantity) for component in body.components],
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=422, detail={"error_type": "invalid_apu", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(x_actor, "apu_updated", concept_code, catalog=catalog)
    return {"concept_code": concept_code, "components": len(body.components)}


@router.put("/rendimientos/{concept_code}")
def update_rendimiento(
    concept_code: str,
    body: RendimientoUpdate,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    known = {row["code"] for row in catalog.load_concepts(include_inactive=True)}
    if concept_code not in known:
        raise HTTPException(
            status_code=404,
            detail={"error_type": "concept_not_found", "code": concept_code},
        )
    catalog.set_rendimiento(concept_code, body.production_rate_per_day)
    _publish_catalog_updated(x_actor, "rendimiento_updated", concept_code, catalog=catalog)
    return {
        "concept_code": concept_code,
        "production_rate_per_day": body.production_rate_per_day,
    }


@router.post("/concepts", status_code=201)
def create_concept(
    body: ConceptCreate,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    try:
        row = catalog.create_concept(
            code=body.code.upper(),
            description=body.description,
            unit=body.unit.upper(),
            phase=body.phase,
            production_rate_per_day=body.production_rate_per_day,
            components=[(c.resource_code, c.quantity) for c in body.components],
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=422, detail={"error_type": "invalid_concept", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(x_actor, "concept_created", body.code.upper(), catalog=catalog)
    return row


@router.put("/concepts/{concept_code}")
def update_concept(
    concept_code: str,
    body: ConceptUpdate,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    try:
        row = catalog.update_concept(
            concept_code,
            description=body.description,
            unit=body.unit,
            phase=body.phase,
            production_rate_per_day=body.production_rate_per_day,
            active=body.active,
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=404 if "no existe" in str(exc) else 422,
            detail={"error_type": "invalid_concept", "message": str(exc)},
        ) from exc
    _publish_catalog_updated(x_actor, "concept_updated", concept_code, catalog=catalog)
    return row


@router.post("/import-prices")
async def import_prices(
    request: Request,
    file: UploadFile,
    x_actor: Annotated[str | None, Header()] = None,
    source: str = "Importación CSV",
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    """Bulk price update from a UTF-8 CSV with columns code,unit_cost (or
    clave,costo_unitario). Only existing insumo codes are updated."""
    require_catalog_admin(request)
    # Read at most the limit plus one byte: a bigger body never lands in memory.
    raw = await file.read(MAX_IMPORT_BYTES + 1)
    if len(raw) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=413,
            detail={"error_type": "import_too_large", "max_bytes": MAX_IMPORT_BYTES},
        )
    filename = (file.filename or "").lower()
    if filename.endswith((".xlsx", ".xlsm")) or raw[:2] == b"PK":
        rows = _rows_from_xlsx(raw)
    else:
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = raw.decode("cp1252")
            except UnicodeDecodeError as exc:
                raise HTTPException(
                    status_code=422,
                    detail={
                        "error_type": "invalid_encoding",
                        "message": "El archivo debe ser CSV UTF-8/CP1252 o XLSX.",
                    },
                ) from exc
        delimiter = ";" if text.count(";") > text.count(",") else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        rows = [
            {(key or "").strip().lower(): (value or "") for key, value in row.items()}
            for row in reader
        ]
    if not rows:
        raise HTTPException(
            status_code=422,
            detail={
                "error_type": "empty_import",
                "message": "El CSV no contiene filas de datos.",
            },
        )
    result = catalog.import_prices(rows, source=source.strip() or "Importación CSV")
    _publish_catalog_updated(
        x_actor, "prices_imported", f"{result['updated']} precios", catalog=catalog
    )
    return result


# Column aliases across OPUS/Neodata/plain exports: the import maps whichever
# pair of code+price headers the workbook uses.
_CODE_HEADERS = ("code", "clave", "código", "codigo", "key")
_COST_HEADERS = (
    "unit_cost", "costo_unitario", "costo", "precio", "precio unitario",
    "p.u.", "pu", "precio_unitario", "precio cotizado", "cotización", "cotizacion",
    "precio cotizado (mxn)",
)


def _rows_from_xlsx(raw: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=422,
            detail={"error_type": "invalid_xlsx", "message": "No se pudo leer el XLSX."},
        ) from exc
    ws = workbook.active
    header_row: list[str] | None = None
    code_index = cost_index = -1
    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(values_only=True):
        values = ["" if v is None else str(v).strip() for v in row]
        if header_row is None:
            lowered = [v.lower() for v in values]
            code_candidates = [i for i, v in enumerate(lowered) if v in _CODE_HEADERS]
            cost_candidates = [i for i, v in enumerate(lowered) if v in _COST_HEADERS]
            if code_candidates and cost_candidates:
                header_row = values
                code_index, cost_index = code_candidates[0], cost_candidates[0]
            continue
        if code_index < len(values) and cost_index < len(values):
            rows.append({"code": values[code_index], "unit_cost": values[cost_index]})
    if header_row is None:
        raise HTTPException(
            status_code=422,
            detail={
                "error_type": "headers_not_found",
                "message": (
                    "No se encontró un encabezado con clave/código y precio; "
                    "exporta desde OPUS/Neodata con esas columnas."
                ),
            },
        )
    return rows


# ------------------------------------------------------------ reference library

class AdoptInput(BaseModel):
    ref_id: int
    # Adopt a price in another unit on purpose; the note must say why.
    force: bool = False
    note: str = Field(default="", max_length=300)


class LaborInput(BaseModel):
    params: FsrParameters
    categories: list[LaborCategory] = Field(min_length=1, max_length=40)


class EquipmentInput(BaseModel):
    description: str | None = Field(default=None, max_length=200)
    params: EquipmentParameters


@router.get("/sources")
def list_reference_sources(
    catalog: CatalogStore = Depends(get_catalog),
    settings: Settings = Depends(get_settings),
) -> dict:
    """Known publications, whether their file is on this server, and whether
    they are imported into the library; then the taller's own catálogos."""
    imported = {row["source_key"]: row for row in catalog.list_sources()}
    known = available_sources(settings.data_dir)
    known_keys = {spec["key"] for spec in known}
    own = [
        {
            "key": row["source_key"], "name": row["name"], "publisher": row["publisher"],
            "region": row["region"], "vigencia": row["vigencia"], "kind": row["kind"],
            "filename": "", "url": "", "available": True, "custom": True, "imported": row,
        }
        for key, row in imported.items() if key not in known_keys
    ]
    return {
        "sources": [{**spec, "imported": imported.get(spec["key"])} for spec in known] + own
    }


@router.post("/sources/custom", status_code=201)
async def import_custom_source(
    request: Request,
    file: UploadFile,
    x_actor: Annotated[str | None, Header()] = None,
    name: str = "",
    vigencia: str = "",
    publisher: str = "",
    kind: Literal["precios_unitarios", "mano_de_obra", "costo_horario"] = "precios_unitarios",
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    """The taller's own catálogo de conceptos (XLSX/CSV with clave,
    descripción, unidad, precio unitario) as a reference source, labeled as
    the taller's, never as a publication.

    ``kind`` dice qué incluye el precio, y no es un detalle: un catálogo de
    destajos trae sólo la mano de obra —la línea sanitaria de 4" vale $61.76
    de mano de obra y $315.59 con el tubo— y adoptar uno creyendo que es el
    precio completo deja el presupuesto corto por el material entero."""
    require_catalog_admin(request)
    # Read at most the limit plus one byte: a bigger body never lands in memory.
    raw = await file.read(MAX_IMPORT_BYTES + 1)
    if len(raw) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=413,
            detail={"error_type": "import_too_large", "max_bytes": MAX_IMPORT_BYTES},
        )
    title = name.strip() or Path(file.filename or "catalogo").stem[:80]
    try:
        rows = parse_concept_workbook(raw, file.filename or "")
    except CustomCatalogError as exc:
        raise HTTPException(
            status_code=422, detail={"error_type": "headers_not_found", "message": str(exc)}
        ) from exc
    key = source_key_for(title)
    count = catalog.import_reference(
        {
            "key": key, "name": title, "publisher": publisher.strip() or "Catálogo propio",
            "region": "MX", "vigencia": vigencia.strip(), "kind": kind, "url": "",
        },
        rows,
        sha256=hashlib.sha256(raw).hexdigest(),
    )
    _publish_catalog_updated(
        x_actor, "reference_imported", f"{title}: {count} renglones", catalog=catalog
    )
    return {"source_key": key, "rows": count, "name": title}


@router.delete("/sources/{source_key}")
def delete_source(
    source_key: str,
    request: Request,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    """Quitar una fuente importada y sus renglones.

    Sin esto, un catálogo importado con el alcance equivocado —un destajo
    marcado como precio unitario completo— se queda para siempre, compitiendo
    en el buscador contra el bueno."""
    require_catalog_admin(request)
    try:
        removed = catalog.delete_source(source_key)
    except ValueError as exc:
        raise HTTPException(
            status_code=409, detail={"error_type": "source_in_use", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(
        x_actor, "reference_removed", f"{removed['name']}: {removed['rows']} renglones",
        catalog=catalog,
    )
    return removed


@router.post("/import-matrices", status_code=201)
async def import_matrices(
    request: Request,
    file: UploadFile,
    x_actor: Annotated[str | None, Header()] = None,
    source: str = "",
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    """Conceptos con sus matrices from an OPUS/Neodata Excel export (or the
    documented Tipo/Clave/Descripción/Unidad/Cantidad/Costo layout)."""
    require_catalog_admin(request)
    # Read at most the limit plus one byte: a bigger body never lands in memory.
    raw = await file.read(MAX_IMPORT_BYTES + 1)
    if len(raw) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=413,
            detail={"error_type": "import_too_large", "max_bytes": MAX_IMPORT_BYTES},
        )
    label = source.strip() or Path(file.filename or "matrices").stem[:80]
    try:
        parse = parse_matrices_workbook(raw, file.filename or "")
    except CustomCatalogError as exc:
        raise HTTPException(
            status_code=422, detail={"error_type": "headers_not_found", "message": str(exc)}
        ) from exc
    result = catalog.import_matrices(parse, label)
    _publish_catalog_updated(
        x_actor, "matrices_imported",
        f"{label}: {result['concepts_created']} nuevos, {result['concepts_updated']} actualizados",
     catalog=catalog)
    return result


@router.post("/sources/{source_key}/import")
def import_reference_source(
    request: Request,
    source_key: str,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
    settings: Settings = Depends(get_settings),
) -> dict:
    require_catalog_admin(request)
    spec = SOURCES.get(source_key)
    if spec is None:
        raise HTTPException(status_code=404, detail={"error_type": "source_not_found"})
    path = sources_dir(settings.data_dir) / spec.filename
    if not path.exists():
        raise HTTPException(
            status_code=409,
            detail={
                "error_type": "source_file_missing",
                "message": f"Descarga {spec.filename} en data/sources antes de importar.",
                "url": spec.url,
            },
        )
    manifest = next(
        (e for e in available_sources(settings.data_dir) if e["key"] == source_key), {}
    )
    count = catalog.import_reference(
        {
            "key": spec.key, "name": spec.name, "publisher": spec.publisher,
            "region": spec.region, "vigencia": spec.vigencia, "kind": spec.kind, "url": spec.url,
        },
        spec.parser(path),
        sha256=manifest.get("sha256"),
    )
    _publish_catalog_updated(
        x_actor, "reference_imported", f"{spec.name}: {count} renglones", catalog=catalog
    )
    return {"source_key": source_key, "rows": count}


# ------------------------------------------------------------ descripciones LOPSRM

@router.get("/concepts/{code}/descripcion")
def concept_long_description(code: str, catalog: CatalogStore = Depends(get_catalog)) -> dict:
    """The LOPSRM-style long description, derived from the concept and its
    matrix — a proposal to paste into the taller's catálogo, never silent."""
    concepts = build_catalog_from_store(catalog.load_concepts(), CostingAssumptions())
    concept = next((c for c in concepts if c.code == code), None)
    if concept is None:
        raise HTTPException(status_code=404, detail={"error_type": "concept_not_found"})
    try:
        apus = build_all_apus(
            concepts, catalog.load_price_book(), templates=catalog.load_templates()
        )
    except ReportGenerationError:
        apus = {}
    return {
        "concept_code": code,
        "description": concept.description,
        "long_description": long_description(concept, apus.get(code)),
        "generated": True,
    }


# ------------------------------------------------------------ vigencia de precios

class IndicesInput(BaseModel):
    source: str = Field(default="", max_length=200)
    values: dict[str, float] = Field(default_factory=dict)


class RollForwardInput(BaseModel):
    codes: list[str] = Field(default_factory=list, max_length=2000)
    status: Literal["vencido", "revisar", "all"] | None = None
    to_month: str | None = Field(default=None, max_length=7)
    dry_run: bool = False


@router.get("/vigencia")
def price_vigencia(catalog: CatalogStore = Depends(get_catalog)) -> dict:
    ages = price_ages(catalog.list_insumos())
    counts = {"vigente": 0, "revisar": 0, "vencido": 0}
    for age in ages:
        counts[age.status] += 1
    return {
        "insumos": [age.__dict__ for age in ages],
        "counts": counts,
        "fresh_months": FRESH_MONTHS,
        "stale_months": STALE_MONTHS,
    }


@router.get("/cotizacion.xlsx")
def cotizacion_request(
    status: Literal["vencido", "revisar", "all"] = "vencido",
    codes: str = "",
    catalog: CatalogStore = Depends(get_catalog),
) -> Response:
    """Solicitud de cotización: the insumos to re-price, with their current
    price and vigencia, and empty columns for the supplier; the answer
    imports through Importar precios (column 'Precio cotizado')."""
    wanted = {c.strip().upper() for c in codes.split(",") if c.strip()}
    ages = [
        a for a in price_ages(catalog.list_insumos())
        if (a.code in wanted if wanted else (status == "all" or a.status == status))
    ]
    content = build_cotizacion_workbook(ages)
    filename = f"solicitud_cotizacion_{now_month()}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/indices")
def get_indices(catalog: CatalogStore = Depends(get_catalog)) -> dict:
    return catalog.load_indices()


@router.put("/indices")
def put_indices(
    request: Request,
    body: IndicesInput,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    require_catalog_admin(request)
    saved = catalog.save_indices(body.source, body.values)
    _publish_catalog_updated(
        x_actor, "indices_saved", f"{len(saved['values'])} meses", catalog=catalog
    )
    return saved


@router.post("/indices/roll-forward")
def roll_forward(
    request: Request,
    body: RollForwardInput,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    """Bring old prices to this month by the taller's index table; each
    updated insumo becomes 'calculado' with the factor as its provenance.
    With ``dry_run`` nothing is written: the response is the preview of
    exactly what a real call with the same body would change."""
    require_catalog_admin(request)
    indices = catalog.load_indices()
    values = indices.get("values") or {}
    if not values:
        raise HTTPException(
            status_code=422,
            detail={
                "error_type": "no_indices", "message": "Captura primero la tabla de índices.",
            },
        )
    to_month = (body.to_month or now_month())[:7]
    wanted = {c.strip().upper() for c in body.codes if c.strip()}
    updated: list[dict] = []
    skipped: list[str] = []
    for age in price_ages(catalog.list_insumos()):
        if wanted and age.code not in wanted:
            continue
        if not wanted and body.status and body.status != "all" and age.status != body.status:
            continue
        if not wanted and not body.status:
            continue
        if not age.vigencia:
            skipped.append(f"{age.code}: sin vigencia, no hay de dónde actualizar")
            continue
        factor = roll_forward_factor(values, age.vigencia, to_month)
        if factor is None:
            skipped.append(f"{age.code}: sin índice para {age.vigencia[:7]}")
            continue
        new_cost = round(age.unit_cost * factor, 2)
        if not body.dry_run:
            catalog.upsert_insumo(
                age.code, unit_cost=new_cost, source_type="calculado",
                source=(
                    f"{age.source or 'precio anterior'} × índice "
                    f"{indices.get('source') or 'del taller'} "
                    f"{age.vigencia[:7]}→{to_month} (factor {factor:.4f})"
                )[:200],
                vigencia=to_month,
            )
        updated.append({
            "code": age.code, "description": age.description, "from": age.unit_cost,
            "to": new_cost, "factor": factor, "vigencia_from": age.vigencia[:7],
        })
    if not body.dry_run:
        _publish_catalog_updated(
            x_actor, "prices_rolled_forward", f"{len(updated)} insumos", catalog=catalog
        )
    return {
        "updated": updated, "skipped": skipped, "to_month": to_month, "dry_run": body.dry_run,
    }


# ------------------------------------------------------------ plantillas & paramétricos

class ParametricRuleInput(BaseModel):
    concept_code: str = Field(min_length=2, max_length=40)
    basis: str = Field(default="m2_construida", max_length=60)
    factor: float = Field(gt=0)
    source: str = Field(default="", max_length=120)
    note: str = Field(default="", max_length=300)


class ParametricRuleUpdate(BaseModel):
    factor: float | None = Field(default=None, gt=0)
    active: bool | None = None
    note: str | None = Field(default=None, max_length=300)


@router.get("/plantillas")
def list_plantillas(catalog: CatalogStore = Depends(get_catalog)) -> dict:
    return {
        "plantillas": catalog.list_plantillas(),
        "rules": catalog.list_parametric_rules(include_inactive=True),
    }


@router.post("/plantillas", status_code=201)
async def create_plantilla(
    request: Request,
    file: UploadFile,
    x_actor: Annotated[str | None, Header()] = None,
    name: str = "",
    tipologia: str = "",
    area_m2: float = 0.0,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    """A past presupuesto (XLSX/CSV with clave, unidad, cantidad, precio) plus
    that project's m² → a plantilla: per-m² rules and the taller's prices."""
    require_catalog_admin(request)
    # Read at most the limit plus one byte: a bigger body never lands in memory.
    raw = await file.read(MAX_IMPORT_BYTES + 1)
    if len(raw) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=413,
            detail={"error_type": "import_too_large", "max_bytes": MAX_IMPORT_BYTES},
        )
    label = name.strip() or Path(file.filename or "plantilla").stem[:80]
    try:
        result = import_plantilla(
            catalog, raw, file.filename or "", name=label, tipologia=tipologia.strip(),
            area_m2=area_m2, actor=clean_actor(x_actor) or "",
        )
    except CustomCatalogError as exc:
        raise HTTPException(
            status_code=422, detail={"error_type": "headers_not_found", "message": str(exc)}
        ) from exc
    except ValueError as exc:
        raise HTTPException(
            status_code=422, detail={"error_type": "invalid_plantilla", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(
        x_actor, "plantilla_imported",
        f"{label}: {result['rules']} reglas, {result['concepts_created']} conceptos nuevos",
     catalog=catalog)
    return result


@router.delete("/plantillas/{key}")
def delete_plantilla(
    request: Request,
    key: str,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    require_catalog_admin(request)
    if not catalog.delete_plantilla(key):
        raise HTTPException(status_code=404, detail={"error_type": "plantilla_not_found"})
    _publish_catalog_updated(x_actor, "plantilla_deleted", key, catalog=catalog)
    return {"deleted": key}


@router.post("/parametrics", status_code=201)
def add_parametric_rule(
    body: ParametricRuleInput,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    try:
        rule = catalog.add_parametric_rule(
            concept_code=body.concept_code.strip().upper(), basis=body.basis.strip(),
            factor=body.factor, source=body.source.strip() or "regla del taller",
            note=body.note.strip(),
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=422, detail={"error_type": "invalid_rule", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(x_actor, "parametric_added", body.concept_code, catalog=catalog)
    return rule


@router.put("/parametrics/{rule_id}")
def update_parametric_rule(
    rule_id: int,
    body: ParametricRuleUpdate,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    try:
        rule = catalog.update_parametric_rule(
            rule_id, factor=body.factor, active=body.active, note=body.note
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=404, detail={"error_type": "rule_not_found", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(x_actor, "parametric_updated", str(rule_id), catalog=catalog)
    return rule


@router.delete("/parametrics/{rule_id}")
def delete_parametric_rule(
    request: Request,
    rule_id: int,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    require_catalog_admin(request)
    if not catalog.delete_parametric_rule(rule_id):
        raise HTTPException(status_code=404, detail={"error_type": "rule_not_found"})
    _publish_catalog_updated(x_actor, "parametric_deleted", str(rule_id), catalog=catalog)
    return {"deleted": rule_id}


# ------------------------------------------------------------ aliases & matching

class AliasInput(BaseModel):
    concept_code: str = Field(min_length=2, max_length=40)
    kind: Literal["reference", "concept"]
    ref_id: int | None = None
    target_code: str | None = Field(default=None, max_length=40)
    note: str = Field(default="", max_length=300)
    project_id: str = Field(default="", max_length=80)
    # Adopt across units on purpose; the note must say why.
    force: bool = False


class BulkAliasInput(BaseModel):
    items: list[AliasInput] = Field(min_length=1, max_length=500)
    project_id: str = Field(default="", max_length=80)


def _candidates(catalog: CatalogStore, source_key: str | None) -> list[Candidate]:
    """Reference rows of every imported source (or the one asked for) and the
    workspace's manual concepts priced by their matrices.

    Antes sólo se buscaba en el catálogo propio del taller, así que los
    tabuladores publicados —5,429 renglones de la CDMX con precio, clave y
    vigencia— se importaban y nadie los veía. El catálogo propio sigue
    ganando cuando dice lo mismo, porque son los precios del taller; lo
    publicado está para lo que el taller todavía no tiene."""
    sources = catalog.list_sources()
    alcance_por_fuente = {
        s["source_key"]: (s.get("kind") or "precios_unitarios") for s in sources
    }
    if source_key:
        keys = [source_key]
    else:
        keys = [s["source_key"] for s in sources]
    candidates: list[Candidate] = []
    if keys or source_key is None:
        rows = catalog.list_reference_rows(keys) if keys else []
        for row in rows:
            candidates.append(
                Candidate(
                    kind="reference", key=str(row["ref_id"]), clave=row["clave"],
                    description=row["description"], unit=row["unit"],
                    price=float(row["price"]), source=row.get("source_name") or "",
                    vigencia=row.get("source_vigencia") or "",
                    phase=row.get("group_description") or "",
                    alcance=alcance_por_fuente.get(
                        row.get("source_key") or "", "precios_unitarios"
                    ),
                )
            )
    if source_key is None:
        concepts = build_catalog_from_store(catalog.load_concepts(), CostingAssumptions())
        try:
            apus = build_all_apus(
                concepts, catalog.load_price_book(), templates=catalog.load_templates()
            )
        except ReportGenerationError:
            apus = {}
        for concept in concepts:
            if concept.rule is not None:
                continue  # only the taller's own (manual/imported) concepts are targets
            apu = apus.get(concept.code)
            if apu is None:
                # Un concepto sin matriz no puede prestarle un precio a nadie:
                # ofrecerlo como candidato es ofrecer el mismo hueco otra vez.
                continue
            candidates.append(
                Candidate(
                    kind="concept", key=concept.code, clave=concept.code,
                    description=concept.description, unit=concept.unit,
                    price=apu.direct_unit_cost, source="matriz del taller",
                    phase=concept.phase,
                )
            )
    return candidates


def _match_payload(match: Match) -> dict:
    c = match.candidate
    return {
        "kind": c.kind, "key": c.key, "ref_id": int(c.key) if c.kind == "reference" else None,
        "target_code": c.key if c.kind == "concept" else None, "clave": c.clave,
        "description": c.description, "unit": c.unit, "price": c.price, "source": c.source,
        "vigencia": c.vigencia, "score": match.score, "reasons": match.reasons,
        "alcance": c.alcance,
        "solo_mano_de_obra": c.alcance == "mano_de_obra",
    }


@router.get("/concepts/{code}/matches")
def concept_matches(
    code: str,
    source_key: str | None = None,
    spec: str | None = None,
    limit: int = 8,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    """Renglones publicados que podrían darle precio a este concepto.

    ``spec`` es el diámetro que el plano declaró para esta línea — «13 mm
    (1/2")» — y cambia el resultado por completo: «canalización con tubo
    conduit» encuentra su precio al 46 %, y con su diámetro al 100 %. Nadie
    publica el precio de un tubo sin decir de cuál."""
    row = next((c for c in catalog.load_concepts() if c["code"] == code), None)
    if row is None:
        raise HTTPException(status_code=404, detail={"error_type": "concept_not_found"})
    # El diámetro entra en la identidad, antes de cualquier «incluye:»: después
    # de esa palabra es alcance, y el alcance no identifica nada.
    identidad, alcance = split_alcance(row["description"])
    consulta = f"{identidad} de {spec.strip()}" if spec and spec.strip() else identidad
    if alcance:
        consulta = f"{consulta}, incluye {alcance}"
    matches = rank(
        consulta, row["unit"], _candidates(catalog, source_key),
        phase=row["phase"], limit=max(1, min(limit, 30)),
    )
    return {
        "concept_code": code, "spec": spec or "",
        "matches": [_match_payload(m) for m in matches],
    }


def _specs_de_los_proyectos(
    request: Request, store: ProjectStore
) -> dict[str, list[str]]:
    """Material y diámetro que los planos del taller declararon, por concepto.

    Una tubería de agua fría sin diámetro no la cotiza nadie, y el diámetro no
    vive en el catálogo: vive en el proyecto que se midió. Aquí se recogen los
    que ya se leyeron, para buscarle precio a lo que el taller construye de
    verdad y no a un concepto abstracto."""
    keep = visible_project_filter(request, store)
    por_concepto: dict[str, set[str]] = {}
    for project_id in store.list_projects():
        if not keep(project_id):
            continue
        try:
            report = store.read_artifact(project_id, "cost_report.json")
        except Exception:  # noqa: BLE001 — un proyecto sin reporte no aporta nada
            continue
        for line in (report.get("boq") or {}).get("lines") or []:
            code = line.get("concept_code")
            descripcion = (line.get("description") or "").strip()
            if code and descripcion:
                por_concepto.setdefault(code, set()).add(descripcion)
    return {code: sorted(v) for code, v in por_concepto.items()}


@router.get("/matches")
def all_matches(
    request: Request,
    source_key: str | None = None,
    min_score: float = 0.5,
    catalog: CatalogStore = Depends(get_catalog),
    store: ProjectStore = Depends(get_store),
) -> dict:
    """Best candidate per concept without an alias: the bulk review list.

    Cuando los planos del taller ya declararon el material y el diámetro de un
    concepto, se busca con ellos: es la diferencia entre buscarle precio a
    «tubería de agua fría» —que no lo tiene— y a «tubería de agua fría de
    cobre de 25 mm (1")», que sí."""
    rate_limit(request, "matches", max_attempts=60, window_seconds=300.0)
    aliases = catalog.load_concept_aliases()
    candidates = _candidates(catalog, source_key)
    try:
        medidas = _specs_de_los_proyectos(request, store)
    except Exception:  # noqa: BLE001 — sin proyectos legibles se busca sin ellas
        medidas = {}
    out = []
    for row in catalog.load_concepts():
        if row["code"] in aliases or not row.get("rule_key"):
            continue
        # Una descripción por cada forma en que el taller la construye: dos
        # diámetros son dos precios, y ofrecer uno solo esconde al otro. Pero
        # si algún proyecto ya dijo la medida, la versión pelona del catálogo
        # sobra — es la misma pregunta con menos información.
        leidas = medidas.get(row["code"]) or []
        con_medida = [v for v in leidas if v.strip() != row["description"].strip()]
        variantes = con_medida or leidas or [row["description"]]
        for variante in variantes:
            best = rank(variante, row["unit"], candidates, phase=row["phase"], limit=1)
            if not best or best[0].score < min_score:
                continue
            out.append({
                "concept_code": row["code"], "description": variante,
                "leido_del_plano": variante != row["description"],
                "unit": row["unit"], "match": _match_payload(best[0]),
            })
    out.sort(key=lambda item: -item["match"]["score"])
    return {"matches": out, "aliases": len(aliases), "candidates": len(candidates)}


@router.get("/aliases")
def list_aliases(catalog: CatalogStore = Depends(get_catalog)) -> dict:
    return {"aliases": catalog.load_concept_aliases()}


def _recompute_project(
    store: ProjectStore, settings: Settings, project_id: str, actor: str
) -> None:
    """Aliases are workspace-wide; the project that asked sees its numbers move now."""
    if not project_id:
        return
    try:
        root = store.get_root(project_id)
    except HTTPException:
        return
    control_dir = root / settings.processed_dir_name
    overrides = load_overrides(control_dir) or CostingOverrides()
    try:
        report = recompute_and_persist(
            store.artifact_root(project_id), control_dir, root / "reports", project_id,
            overrides,
            catalog_store=store_for_project(settings, project_id),
        )
    except ReportGenerationError:
        return
    BUS.publish(
        "costing_updated",
        project_id=project_id,
        actor=actor,
        data={
            "version": overrides.version, "direct_cost": report.boq.direct_cost_total,
            "grand_total": report.integration.grand_total, "review_action": "alias",
        },
    )


def _forced_without_reason(force: bool, note: str) -> None:
    if force and not note.strip():
        raise HTTPException(
            status_code=422,
            detail={
                "error_type": "reason_required",
                "message": "Forzar una unidad distinta requiere decir por qué.",
            },
        )


def _unit_mismatch(exc: UnitMismatch) -> HTTPException:
    return HTTPException(
        status_code=422,
        detail={
            "error_type": "unit_mismatch", "message": str(exc), "code": exc.code,
            "unit": exc.own_unit, "reference_unit": exc.other_unit,
        },
    )


def _set_alias(catalog: CatalogStore, item: AliasInput, actor: str, project_id: str) -> dict:
    _forced_without_reason(item.force, item.note)
    try:
        return catalog.set_concept_alias(
            item.concept_code.strip().upper(), kind=item.kind, ref_id=item.ref_id,
            target_code=(item.target_code or "").strip().upper() or None, actor=actor,
            note=item.note.strip(), project_id=project_id or item.project_id,
            force=item.force,
        )
    except UnitMismatch as exc:
        raise _unit_mismatch(exc) from exc
    except ValueError as exc:
        raise HTTPException(
            status_code=422, detail={"error_type": "invalid_alias", "message": str(exc)}
        ) from exc


@router.post("/aliases", status_code=201)
def set_alias(
    body: AliasInput,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
    store: ProjectStore = Depends(get_store),
    settings: Settings = Depends(get_settings),
) -> dict:
    actor = clean_actor(x_actor) or ""
    alias = _set_alias(catalog, body, actor, body.project_id)
    _publish_catalog_updated(
        x_actor, "alias_set", f"{body.concept_code} → {alias['clave']}", catalog=catalog
    )
    _recompute_project(store, settings, body.project_id, actor)
    return alias


@router.post("/aliases/bulk", status_code=201)
def set_aliases_bulk(
    body: BulkAliasInput,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
    store: ProjectStore = Depends(get_store),
    settings: Settings = Depends(get_settings),
) -> dict:
    actor = clean_actor(x_actor) or ""
    saved = [_set_alias(catalog, item, actor, body.project_id) for item in body.items]
    _publish_catalog_updated(
        x_actor, "aliases_set", f"{len(saved)} conceptos del taller", catalog=catalog
    )
    _recompute_project(store, settings, body.project_id, actor)
    return {"aliases": saved}


@router.delete("/aliases/{code}")
def clear_alias(
    code: str,
    project_id: str = "",
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
    store: ProjectStore = Depends(get_store),
    settings: Settings = Depends(get_settings),
) -> dict:
    if not catalog.clear_concept_alias(code.upper()):
        raise HTTPException(status_code=404, detail={"error_type": "alias_not_found"})
    actor = clean_actor(x_actor) or ""
    _publish_catalog_updated(x_actor, "alias_cleared", code, catalog=catalog)
    _recompute_project(store, settings, project_id, actor)
    return {"cleared": code.upper()}


@router.get("/reference")
def search_reference(
    q: str = "",
    source: str | None = None,
    limit: int = 50,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    if len(q.strip()) < 2:
        return {"rows": []}
    return {"rows": catalog.search_reference(q, source_key=source, limit=limit)}


@router.post("/insumos/{code}/adopt")
def adopt_reference_price(
    code: str,
    body: AdoptInput,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    if code not in catalog.load_price_book():
        raise HTTPException(status_code=404, detail={"error_type": "insumo_not_found"})
    _forced_without_reason(body.force, body.note)
    try:
        row = catalog.adopt_reference(code, body.ref_id, force=body.force)
    except UnitMismatch as exc:
        raise _unit_mismatch(exc) from exc
    except ValueError as exc:
        raise HTTPException(
            status_code=404, detail={"error_type": "reference_not_found", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(x_actor, "insumo_updated", code, catalog=catalog)
    return row


@router.post("/concepts/{code}/adopt")
def adopt_concept_price(
    code: str,
    body: AdoptInput,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    """The concept's P.U. becomes the reference row's price (catálogo propio
    or publication), replacing its matrix until cleared."""
    _forced_without_reason(body.force, body.note)
    try:
        row = catalog.adopt_concept_reference(code, body.ref_id, force=body.force)
    except UnitMismatch as exc:
        raise _unit_mismatch(exc) from exc
    except ValueError as exc:
        raise HTTPException(
            status_code=404, detail={"error_type": "not_found", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(x_actor, "concept_updated", code, catalog=catalog)
    return row


@router.delete("/concepts/{code}/price")
def clear_concept_price(
    code: str,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    try:
        row = catalog.clear_concept_price(code)
    except ValueError as exc:
        raise HTTPException(
            status_code=404, detail={"error_type": "concept_not_found", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(x_actor, "concept_updated", code, catalog=catalog)
    return row


class InventoryMappingInput(BaseModel):
    kind: Literal["block", "layer", "tag", "area"]
    pattern: str = Field(min_length=1, max_length=200)
    concept_code: str = Field(min_length=1, max_length=40)
    factor: float = Field(default=1.0, gt=0)


@router.get("/inventory-mappings")
def list_inventory_mappings(catalog: CatalogStore = Depends(get_catalog)) -> dict:
    """Symbol/layer → concept rules: how a levantamiento count becomes a quantity."""
    return {"mappings": catalog.list_inventory_mappings()}


@router.post("/inventory-mappings", status_code=201)
def add_inventory_mapping(
    body: InventoryMappingInput,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    try:
        row = catalog.add_inventory_mapping(
            kind=body.kind, pattern=body.pattern, concept_code=body.concept_code,
            factor=body.factor,
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=422, detail={"error_type": "invalid_mapping", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(
        x_actor, "inventory_mapping", f"{body.pattern} → {body.concept_code}"
    , catalog=catalog)
    return row


@router.delete("/inventory-mappings/{mapping_id}")
def delete_inventory_mapping(
    mapping_id: int,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    if not catalog.delete_inventory_mapping(mapping_id):
        raise HTTPException(status_code=404, detail={"error_type": "mapping_not_found"})
    _publish_catalog_updated(
        x_actor, "inventory_mapping", f"#{mapping_id} eliminada", catalog=catalog
    )
    return {"deleted": mapping_id}


@router.get("/labor")
def get_labor(catalog: CatalogStore = Depends(get_catalog)) -> dict:
    return labor_state(catalog)


@router.get("/labor/presets")
def get_labor_presets() -> dict:
    """Regional presets: minimum-wage zone and ISN per state, with sources."""
    return {"presets": [p.model_dump() for p in REGION_PRESETS]}


@router.post("/labor/presets/{key}")
def apply_labor_preset(
    request: Request,
    key: str,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    """Apply a state's ISN and minimum wage to the saved parameters and
    categories, and reprice labor."""
    require_catalog_admin(request)
    preset = preset_by_key(key)
    if preset is None:
        raise HTTPException(status_code=404, detail={"error_type": "preset_not_found"})
    state = labor_state(catalog)
    params = FsrParameters.model_validate(state["params"])
    categories = [
        LaborCategory.model_validate({k: v for k, v in c.items() if k != "breakdown"})
        for c in state["categories"]
    ]
    params, categories = apply_preset(params, categories, preset)
    applied = apply_labor(catalog, params, categories)
    _publish_catalog_updated(x_actor, "labor_preset_applied", preset.label, catalog=catalog)
    return {"preset": preset.model_dump(), "applied": applied, **labor_state(catalog)}


@router.post("/labor/preview")
def preview_labor(body: LaborInput, catalog: CatalogStore = Depends(get_catalog)) -> dict:
    """What "Aplicar salario real" would write, next to what the catálogo
    prices each category at today. Nothing is saved."""
    current = {row["code"]: row for row in catalog.list_insumos()}
    rows = []
    for category in body.categories:
        breakdown = compute_fsr(category.salario_nominal, body.params)
        before = current.get(category.code)
        rows.append({
            "code": category.code,
            "description": category.description,
            "salario_nominal": category.salario_nominal,
            "fsr": breakdown.fsr,
            "from": before["unit_cost"] if before else None,
            "to": breakdown.salario_real,
        })
    return {"rows": rows, "vigencia": now_month()}


@router.put("/labor")
def put_labor(
    request: Request,
    body: LaborInput,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    require_catalog_admin(request)
    applied = apply_labor(catalog, body.params, body.categories)
    _publish_catalog_updated(
        x_actor, "labor_applied", f"{len(applied)} categorías", catalog=catalog
    )
    return {"applied": applied, **labor_state(catalog)}


@router.get("/equipment/{code}")
def get_equipment(code: str, catalog: CatalogStore = Depends(get_catalog)) -> dict:
    analysis = catalog.get_analysis(code)
    if analysis and analysis["kind"] == "costo_horario":
        params = EquipmentParameters.model_validate(analysis["params"])
        return {"code": code, "params": params.model_dump(),
                "breakdown": compute_costo_horario(params).model_dump(), "saved": True}
    return {"code": code, "params": None, "breakdown": None, "saved": False}


@router.put("/equipment/{code}")
def put_equipment(
    request: Request,
    code: str,
    body: EquipmentInput,
    x_actor: Annotated[str | None, Header()] = None,
    catalog: CatalogStore = Depends(get_catalog),
) -> dict:
    require_catalog_admin(request)
    try:
        row = apply_equipment(catalog, code, body.description, body.params)
    except ValueError as exc:
        raise HTTPException(
            status_code=422, detail={"error_type": "invalid_insumo", "message": str(exc)}
        ) from exc
    _publish_catalog_updated(x_actor, "insumo_updated", code, catalog=catalog)
    return row
