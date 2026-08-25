"""Los convenios cambian lo contratado, y la estimación que sigue debe saberlo.

La lógica de cada módulo ya se prueba aparte; lo que se prueba aquí es la
costura, que es donde un contrato modificado se queda a medias: si la
estimación siguiente sigue leyendo el catálogo original, va a marcar como
excedido algo que ya se convino, y el aviso que sí importa se pierde entre los
que ya no."""

import json

import pytest
from klave_engine.common import config as config_module


@pytest.fixture
def client(data_dir, monkeypatch):
    from fastapi.testclient import TestClient

    from apps.api.main import create_app

    monkeypatch.setenv("KLAVE_USERS_DATABASE_URL", "postgresql://nobody@127.0.0.1:1/none")
    config_module.get_settings.cache_clear()
    return TestClient(create_app())


def _proyecto(data_dir) -> str:
    """Un proyecto mínimo en disco: los artefactos de obra no salen del plano."""
    pid = "obra-test"
    root = data_dir / "projects" / pid
    (root / "processed").mkdir(parents=True, exist_ok=True)
    (root / "processed" / "project_manifest.json").write_text(
        json.dumps({
            "project_id": pid, "project_name": "Obra de prueba",
            "root_path": str(root), "drawings": [],
        })
    )
    (data_dir / "projects_registry.json").write_text(json.dumps({pid: str(root)}))
    return pid


def _escribir(data_dir, pid: str, nombre: str, datos: object) -> None:
    (data_dir / "projects" / pid / "processed" / nombre).write_text(json.dumps(datos))


ESTIMACION_1 = {
    "numero": 1,
    "periodo_inicio": "2026-01-01",
    "periodo_fin": "2026-01-31",
    "monto_contrato": 50_000.0,
    "anticipo_pct": 30.0,
    "retencion_pct": 5.0,
    "renglones": [
        {
            "clave": "OP-001", "description": "Muro de tabique", "unit": "m2",
            "unit_price": 500.0, "quantity_period": 100.0, "quantity_previous": 0.0,
            "quantity_contract": 100.0,
        }
    ],
}


def test_la_estimacion_siguiente_lee_el_catalogo_ya_convenido(client, data_dir):
    pid = _proyecto(data_dir)
    _escribir(data_dir, pid, "estimaciones.json", [ESTIMACION_1])
    _escribir(data_dir, pid, "catalogo_convocante.json",
              {"renglones": [{"clave": "OP-001", "quantity": 100.0, "unit_price": 500.0}]})

    # Sin convenio: lo contratado sigue siendo 100.
    sin = client.post(f"/projects/{pid}/estimaciones/siguiente",
                      params={"inicio": "2026-02-01", "fin": "2026-02-28"})
    assert sin.status_code == 200, sin.text
    assert sin.json()["estimacion"]["renglones"][0]["quantity_contract"] == 100.0

    conv = client.put(f"/projects/{pid}/convenios/1", json={"convenio": {
        "numero": 1, "fecha": "2026-02-05", "motivo": "Ampliación de fachada",
        "renglones": [{
            "clave": "OP-001", "description": "Muro de tabique", "unit": "m2",
            "unit_price": 500.0, "quantity": 120.0, "quantity_anterior": 100.0,
        }],
    }})
    assert conv.status_code == 200, conv.text

    con = client.post(f"/projects/{pid}/estimaciones/siguiente",
                      params={"inicio": "2026-02-01", "fin": "2026-02-28"})
    renglon = con.json()["estimacion"]["renglones"][0]
    assert renglon["quantity_contract"] == 120.0
    # Y el monto del contrato crece con lo convenido, para amortizar bien:
    # 50 000 del catálogo más 10 000 del convenio.
    assert con.json()["estimacion"]["monto_contrato"] == 60_000.0


def test_un_concepto_que_entra_por_convenio_aparece_en_la_estimacion(client, data_dir):
    pid = _proyecto(data_dir)
    _escribir(data_dir, pid, "estimaciones.json", [ESTIMACION_1])
    client.put(f"/projects/{pid}/convenios/1", json={"convenio": {
        "numero": 1, "fecha": "2026-02-05", "motivo": "Concepto no previsto",
        "renglones": [{
            "clave": "OP-050", "description": "Impermeabilizante", "unit": "m2",
            "unit_price": 200.0, "quantity": 80.0, "quantity_anterior": 0.0,
        }],
    }})
    est = client.post(f"/projects/{pid}/estimaciones/siguiente",
                      params={"inicio": "2026-02-01", "fin": "2026-02-28"}).json()

    nuevo = [r for r in est["estimacion"]["renglones"] if r["clave"] == "OP-050"]
    assert len(nuevo) == 1
    assert nuevo[0]["quantity_contract"] == 80.0
    # Entra en cero: existe en el contrato, pero nadie lo ha medido todavía.
    assert nuevo[0]["quantity_period"] == 0.0


def test_el_techo_del_articulo_59_sale_en_el_estado(client, data_dir):
    pid = _proyecto(data_dir)
    _escribir(data_dir, pid, "catalogo_convocante.json",
              {"renglones": [{"clave": "OP-001", "quantity": 100.0, "unit_price": 500.0}]})
    client.put(f"/projects/{pid}/convenios/1", json={"convenio": {
        "numero": 1, "fecha": "2026-02-05",
        "renglones": [{
            "clave": "OP-001", "description": "Muro", "unit": "m2",
            "unit_price": 500.0, "quantity": 140.0, "quantity_anterior": 100.0,
        }],
    }})
    estado = client.get(f"/projects/{pid}/convenios").json()["estado"]

    assert estado["monto_original"] == 50_000.0
    assert estado["monto_convenido"] == 20_000.0
    assert estado["monto_pct"] == 40.0
    assert estado["rebasa_techo"] is True
    assert any("art. 59" in a for a in estado["avisos"])


def test_el_borrador_de_convenio_no_se_inventa_cuando_no_hay_excedente(client, data_dir):
    pid = _proyecto(data_dir)
    _escribir(data_dir, pid, "estimaciones.json", [ESTIMACION_1])
    r = client.post(f"/projects/{pid}/convenios/desde-estimacion/1",
                    params={"fecha": "2026-02-05"})
    assert r.status_code == 409
    assert r.json()["detail"]["error_type"] == "sin_excedentes"


def test_el_borrador_toma_lo_ejecutado_como_nueva_cantidad(client, data_dir):
    pid = _proyecto(data_dir)
    excedida = json.loads(json.dumps(ESTIMACION_1))
    excedida["renglones"][0]["quantity_period"] = 130.0
    _escribir(data_dir, pid, "estimaciones.json", [excedida])

    conv = client.post(f"/projects/{pid}/convenios/desde-estimacion/1",
                       params={"fecha": "2026-02-05"}).json()["convenio"]
    assert conv["renglones"][0]["quantity"] == 130.0
    assert conv["renglones"][0]["quantity_anterior"] == 100.0
    assert conv["motivo"] == ""


def test_el_finiquito_se_precarga_de_las_estimaciones_capturadas(client, data_dir):
    pid = _proyecto(data_dir)
    _escribir(data_dir, pid, "estimaciones.json", [ESTIMACION_1])
    _escribir(data_dir, pid, "catalogo_convocante.json",
              {"renglones": [{"clave": "OP-001", "quantity": 100.0, "unit_price": 500.0}]})

    cuerpo = client.get(f"/projects/{pid}/finiquito").json()
    assert cuerpo["guardado"] is False
    fin = cuerpo["finiquito"]
    assert fin["ejecutado"] == 50_000.0
    assert fin["retenciones_aplicadas"] == 2_500.0
    # Lo que el motor no puede saber nace en cero y a la vista.
    assert fin["dias_atraso"] == 0
    assert fin["pena_pct_diario"] == 0.0
    assert any("garantía" in s["concepto"] for s in cuerpo["resumen"]["saldos"])
    # Los derivados viajan calculados: la pantalla no debe volver a deducir el
    # signo por su cuenta.
    assert cuerpo["resumen"]["a_favor_de"] == "contratista"
    assert cuerpo["resumen"]["saldos"][0]["a_favor"] == "contratista"


def test_borrar_un_convenio_que_no_existe_es_404(client, data_dir):
    pid = _proyecto(data_dir)
    assert client.delete(f"/projects/{pid}/convenios/7").status_code == 404


def test_sin_catalogo_cargado_el_convenio_no_borra_el_monto_capturado(client, data_dir):
    """No todos los proyectos cargaron el catálogo de la convocante.

    Sin ese respaldo manda lo que se capturó a mano: tomar cero de un catálogo
    ausente dejaría el contrato por los suelos y la amortización del anticipo
    saldría mal, sin que nadie lo notara."""
    pid = _proyecto(data_dir)
    _escribir(data_dir, pid, "estimaciones.json", [ESTIMACION_1])
    client.put(f"/projects/{pid}/convenios/1", json={"convenio": {
        "numero": 1, "fecha": "2026-02-05",
        "renglones": [{
            "clave": "OP-001", "description": "Muro", "unit": "m2",
            "unit_price": 500.0, "quantity": 120.0, "quantity_anterior": 100.0,
        }],
    }})
    est = client.post(f"/projects/{pid}/estimaciones/siguiente",
                      params={"inicio": "2026-02-01", "fin": "2026-02-28"}).json()
    assert est["estimacion"]["monto_contrato"] == 60_000.0


def test_la_tabla_de_unidades_no_vive_bajo_projects(client):
    """GET /projects/{project_id} se la tragaría como si fuera un id de proyecto."""
    r = client.get("/medidas/unidades-generador")
    assert r.status_code == 200, r.text
    unidades = r.json()["unidades"]
    assert unidades["m2"] == ["largo", "ancho"]
    assert unidades["m3"] == ["largo", "ancho", "alto"]
    assert unidades["pza"] == []


def test_la_estimacion_se_exporta_con_su_generador_en_el_mismo_archivo(client, data_dir):
    """Separarlos es la forma más fácil de que uno se quede en el escritorio."""
    pid = _proyecto(data_dir)
    con_generador = json.loads(json.dumps(ESTIMACION_1))
    con_generador["renglones"][0]["generador"] = [
        {"ubicacion": "Eje A-3", "veces": 1, "largo": 20.0, "ancho": 5.0,
         "alto": None, "medida_directa": None, "nota": ""}
    ]
    _escribir(data_dir, pid, "estimaciones.json", [con_generador])

    r = client.get(f"/projects/{pid}/estimaciones/1/export.xlsx")
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]
    assert "estimacion_1" in r.headers["content-disposition"]

    import io

    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(r.content)).active
    texto = "\n".join(
        str(c.value) for fila in ws.iter_rows() for c in fila if c.value is not None
    )
    assert "LÍQUIDO A PAGAR" in texto
    assert "Números generadores" in texto
    assert "Eje A-3" in texto
    # La fórmula viaja escrita: quien revisa rehace la cuenta, no la cree.
    assert "20 × 5" in texto


def test_exportar_una_estimacion_que_no_existe_es_404(client, data_dir):
    pid = _proyecto(data_dir)
    assert client.get(f"/projects/{pid}/estimaciones/9/export.xlsx").status_code == 404


def test_preparar_el_ajuste_carga_lo_pendiente_de_lo_ya_capturado(client, data_dir):
    """Volver a teclearlo es donde se cuelan cantidades que no cuadran."""
    pid = _proyecto(data_dir)
    segunda = json.loads(json.dumps(ESTIMACION_1))
    segunda["numero"] = 2
    segunda["renglones"][0]["quantity_previous"] = 100.0
    segunda["renglones"][0]["quantity_period"] = 40.0
    _escribir(data_dir, pid, "estimaciones.json", [ESTIMACION_1, segunda])

    sol = client.post(f"/projects/{pid}/ajustes/preparar",
                      params={"periodo_base": "2026-01", "periodo_ajuste": "2026-07"})
    assert sol.status_code == 200, sol.text
    renglon = sol.json()["solicitud"]["renglones"][0]
    # El acumulado de la última estimación, no el de la primera.
    assert renglon["quantity_executed"] == 140.0
    assert renglon["quantity_contract"] == 100.0
    # Sin índice no hay factor, y lo dice en vez de aproximar.
    assert sol.json()["resumen"]["factor"] is None
    assert sol.json()["resumen"]["calculable"] is False


def test_el_ajuste_preparado_respeta_el_catalogo_ya_convenido(client, data_dir):
    pid = _proyecto(data_dir)
    _escribir(data_dir, pid, "estimaciones.json", [ESTIMACION_1])
    client.put(f"/projects/{pid}/convenios/1", json={"convenio": {
        "numero": 1, "fecha": "2026-02-05",
        "renglones": [{
            "clave": "OP-001", "description": "Muro", "unit": "m2",
            "unit_price": 500.0, "quantity": 130.0, "quantity_anterior": 100.0,
        }],
    }})
    sol = client.post(f"/projects/{pid}/ajustes/preparar").json()
    assert sol["solicitud"]["renglones"][0]["quantity_contract"] == 130.0


def test_guardar_un_ajuste_con_indice_devuelve_su_factor(client, data_dir):
    pid = _proyecto(data_dir)
    r = client.put(f"/projects/{pid}/ajustes/1", json={"solicitud": {
        "numero": 1, "periodo_base": "2026-01", "periodo_ajuste": "2026-07",
        "indice": {
            "nombre": "INPP construcción", "fuente": "INEGI",
            "publicacion": "captura de prueba",
            "valores": {"2026-01": 100.0, "2026-07": 112.0},
        },
        "renglones": [{
            "clave": "OP-001", "description": "Muro", "unit": "m2", "unit_price": 500.0,
            "quantity_contract": 1000.0, "quantity_executed": 400.0,
        }],
    }})
    assert r.status_code == 200, r.text
    resumen = r.json()["resumen"]
    assert resumen["factor"] == 1.12
    assert resumen["importe_ajuste"] == 36_000.0
    assert client.get(f"/projects/{pid}/ajustes").json()["ajustes"][0]["resumen"]["factor"] == 1.12


NOTA_APERTURA = {
    "numero": 1, "fecha": "2026-01-15", "tipo": "apertura", "parte": "contratante",
    "autor": "Ing. Diego Gaytán", "cargo": "Residente de obra",
    "texto": "Se abre la bitácora del contrato OP-2026-014.",
    "referencia": None, "asentada_en": "",
}


def test_la_bitacora_no_expone_manera_de_editar_ni_de_borrar(client, data_dir):
    """La ausencia es la garantía: si se pudiera corregir, no probaría nada."""
    pid = _proyecto(data_dir)
    rutas = {
        (r.path, m)
        for r in client.app.routes
        for m in getattr(r, "methods", set())
        if "bitacora" in getattr(r, "path", "")
    }
    metodos = {m for _, m in rutas}
    assert metodos <= {"GET", "POST", "HEAD"}
    assert "PUT" not in metodos and "DELETE" not in metodos and "PATCH" not in metodos


def test_asentar_pone_la_hora_del_servidor_no_la_del_navegador(client, data_dir):
    """Si la pusiera el cliente, bastaría mover el reloj para fechar una nota ayer."""
    pid = _proyecto(data_dir)
    mentira = dict(NOTA_APERTURA, asentada_en="1999-01-01T00:00:00+00:00")
    r = client.post(f"/projects/{pid}/bitacora", json={"nota": mentira})
    assert r.status_code == 201, r.text
    assert not r.json()["nota"]["asentada_en"].startswith("1999")


def test_una_nota_repetida_se_rechaza_con_su_razon(client, data_dir):
    pid = _proyecto(data_dir)
    client.post(f"/projects/{pid}/bitacora", json={"nota": NOTA_APERTURA})
    r = client.post(f"/projects/{pid}/bitacora", json={"nota": NOTA_APERTURA})
    assert r.status_code == 409
    assert r.json()["detail"]["error_type"] == "bitacora_rechaza"
    assert "no se reescribe" in r.json()["detail"]["message"]


def test_la_bitacora_se_lee_de_corrido_con_su_estado(client, data_dir):
    pid = _proyecto(data_dir)
    client.post(f"/projects/{pid}/bitacora", json={"nota": NOTA_APERTURA})
    client.post(f"/projects/{pid}/bitacora", json={"nota": dict(
        NOTA_APERTURA, numero=2, tipo="ordinaria", parte="contratista",
        autor="Ing. Ana Ruiz", cargo="Superintendente",
        texto="Se solicita definición de niveles en el eje 4.",
    )})
    cuerpo = client.get(f"/projects/{pid}/bitacora").json()
    assert [n["numero"] for n in cuerpo["notas"]] == [1, 2]
    assert cuerpo["estado"]["abierta"] is True
    assert cuerpo["estado"]["cerrada"] is False
    assert cuerpo["estado"]["siguiente_numero"] == 3
    assert cuerpo["estado"]["por_parte"] == {"contratante": 1, "contratista": 1}
