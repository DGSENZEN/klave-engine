"""Without a reliable unit the engine measures in drawing units and prices
nothing — on screen and in every export — while the detectors still use
thresholds that fit the drawing's extent."""

import io

from klave_engine.costing.exports import (
    IVA_PCT,
    MONEY_FORMAT,
    SIN_UNIDADES,
    UNPRICED,
    blocked_banner,
    build_apus_workbook,
    build_explosion_workbook,
    build_presupuesto_workbook,
)
from klave_engine.costing.models import CostingConfig
from klave_engine.costing.report import generate_cost_report
from klave_engine.costing.reviews import ProjectReviews
from klave_engine.detection.results import DetectionType, make_detection
from klave_engine.detection.suite import DetectorSuiteConfig
from klave_engine.detection.taxonomy import classify_family
from klave_engine.dxf.units import DrawingUnits
from openpyxl import load_workbook

from tests.precios import LIBRO


def _wall(det_id, length):
    det = make_detection(
        det_id, DetectionType.wall, det_id, (0, 0, length, 0.15), 0.9, [], "m", [],
        {"estimated_length": length, "estimated_thickness": 0.15, "wall_kind": "block"},
    )
    det.family = classify_family(det).value
    return det


def test_unknown_units_mean_no_price_anywhere():
    units = DrawingUnits(unit="drawing_units", source="unknown", confidence=0.0)
    report = generate_cost_report("p", [_wall("w1", 1000.0)], units, CostingConfig(), None, None,
        price_book=LIBRO,
    )
    assert not report.boq.units_reliable
    assert report.boq.lines and all(ln.unpriced and ln.amount == 0.0 for ln in report.boq.lines)
    assert report.boq.direct_cost_total == 0.0 and report.integration.grand_total == 0.0
    assert any(w.startswith("SIN UNIDADES") for w in report.boq.warnings)
    for fmt in ("klave", "opus", "licitacion"):
        content = build_presupuesto_workbook(report, [], ProjectReviews(), "Obra", None, fmt=fmt)
        cells = [
            c.value for ws in load_workbook(io.BytesIO(content)).worksheets
            for row in ws.iter_rows() for c in row
        ]
        assert SIN_UNIDADES in cells, fmt


def _labeled_value(ws, label: str, label_col: int, value_col: int):
    """The value beside a summary label, e.g. the amount next to 'TOTAL'.

    Scoped lookups by label+column rather than "is this number anywhere in
    the workbook", because a workbook has many sheets (Programa, Flujo...)
    that may legitimately echo a running total elsewhere — checking the
    exact labeled cell is what actually proves a specific site was gated.
    """
    for row in ws.iter_rows():
        if row[label_col - 1].value == label:
            return row[value_col - 1].value
    raise AssertionError(f"no row labeled {label!r} in sheet {ws.title!r}")


def test_a_priced_but_unconfirmed_report_shows_no_total_in_the_workbook():
    """units_reliable=True here — the engine priced this run for real,
    unlike the fully-unknown-unit case above where grand_total is already
    zero at the source (so asserting its absence there would hold even with
    no gate at all). Confidence sits under CONFIDENCE_FIRM (0.3 < 0.7) with
    no human confirmation, so resolve_money_state still says "blocked" —
    this is the shape that actually proves the three summary cells stop
    leaking a real, nonzero total under their own SIN UNIDADES/SIN VERIFICAR
    banner, in every one of the three functions Task 3 threads money_state
    through."""
    units = DrawingUnits(unit="m", source="dxf_header", confidence=0.3)
    report = generate_cost_report(
        "p", [_wall("w1", 10.0)], units, CostingConfig(), None, None, price_book=LIBRO,
    )
    assert report.boq.units_reliable  # the engine priced it — a real total exists to hide
    assert report.integration.grand_total > 0

    klave = load_workbook(io.BytesIO(
        build_presupuesto_workbook(report, [], ProjectReviews(), "Obra", None, fmt="klave")
    ))
    assert _labeled_value(klave["Carátula"], "Total con contingencia", 1, 2) == UNPRICED
    assert _labeled_value(klave["Presupuesto"], "TOTAL", 5, 6) == UNPRICED

    licitacion = load_workbook(io.BytesIO(
        build_presupuesto_workbook(report, [], ProjectReviews(), "Obra", None, fmt="licitacion")
    ))
    ws = licitacion["Catálogo de conceptos"]
    assert _labeled_value(ws, "SUBTOTAL", 3, 8) == UNPRICED
    assert _labeled_value(ws, f"I.V.A. {IVA_PCT:.0f} %", 3, 8) == UNPRICED
    assert _labeled_value(ws, "TOTAL", 3, 8) == UNPRICED
    # The spelled-out total is the same number in another alphabet: blanking
    # the digits while this cell still reads "...con letra: Setecientos
    # ses.." below them would be the fix undoing itself one row down.
    assert not any(
        isinstance(c.value, str) and c.value.startswith("Importe total con letra")
        for row in ws.iter_rows() for c in row
    )


def _doubtful_report():
    """units_reliable=True and a real, nonzero total — blocked only because
    the reading is 30 % confident and nobody has confirmed it. The shape the
    banner used to lie about, and the shape a leaked addend is visible in."""
    units = DrawingUnits(unit="m", source="dxf_header", confidence=0.3)
    return generate_cost_report(
        "p", [_wall("w1", 10.0)], units, CostingConfig(), None, None, price_book=LIBRO,
    )


def _money_cells(workbook):
    """Every cell the workbook itself formats as pesos, across all sheets.

    Asserting on the format rather than on a list of known labels is what
    makes this survive a new money row: any cell a future edit formats as
    MONEY_FORMAT is caught by the same assertion, with no test to update."""
    return [
        (ws.title, c.coordinate, c.value)
        for ws in workbook.worksheets
        for row in ws.iter_rows()
        for c in row
        if c.number_format == MONEY_FORMAT and c.value is not None
    ]


def test_the_banner_states_the_cause_that_is_actually_true():
    """``blocked`` has three causes and SIN_UNIDADES describes exactly one.

    Emitting it for all three put a sentence reading "la unidad del plano no
    es confiable: cantidades en unidades de dibujo" one row under a carátula
    row reading "Unidades del plano: m (30 % de confianza)" — and every report
    written before this branch carries no money_basis at all, so every legacy
    project's catálogo de conceptos asserted the same false thing in a
    document delivered under LOPSRM art. 45."""
    unreliable = generate_cost_report(
        "p", [_wall("w1", 1000.0)],
        DrawingUnits(unit="drawing_units", source="unknown", confidence=0.0),
        CostingConfig(), None, None, price_book=LIBRO,
    )
    doubtful = _doubtful_report()
    legacy = _doubtful_report()
    legacy.money_basis = None  # a run written before the verdict existed

    # (1) the unit really is unreliable: the sentence that says so is right.
    assert blocked_banner(unreliable.money_basis) == SIN_UNIDADES
    # (2) the unit was read as metres — saying otherwise is the lie.
    doubtful_banner = blocked_banner(doubtful.money_basis)
    assert "unidades de dibujo" not in doubtful_banner
    assert "30 % de confianza" in doubtful_banner.replace("30% de confianza", "30 % de confianza")
    assert "dxf_header" in doubtful_banner
    # (3) no verdict at all is its own cause, and it is not the unit's fault.
    legacy_banner = blocked_banner(legacy.money_basis)
    assert "unidades de dibujo" not in legacy_banner
    assert "corrida anterior sin veredicto de unidades" in legacy_banner
    # All three still say what the reader has to do, and none of them claims
    # the cantidades are unusable — they are the measurements.
    for banner in (doubtful_banner, legacy_banner):
        assert "Klave" in banner and "cantidades" in banner.lower()


def test_a_blocked_workbook_prints_no_peso_figure_at_all(tmp_path):
    """Gating the grand total published its own addends: the presupuesto
    printed "Precio de venta 29,689.18" and "Contingencia 1,484.46" two rows
    above a withheld TOTAL, and the licitación printed three partida
    subtotals above a withheld SUBTOTAL. What is gated is money.

    Checked by format, not by label: every cell the workbook itself marks as
    pesos must hold the sentinel, on every sheet — carátula, presupuesto,
    APUs, explosión, los cuatro programas de erogaciones y el flujo."""
    report = _doubtful_report()
    assert report.integration.grand_total > 0  # a real total exists to leak
    reviews = ProjectReviews()

    for fmt in ("klave", "opus", "neodata", "licitacion", "licitacion_larga"):
        workbook = load_workbook(io.BytesIO(
            build_presupuesto_workbook(report, [], reviews, "Obra", None, fmt=fmt)
        ))
        cells = _money_cells(workbook)
        # Not vacuous: the peso cells are still there, they just say so.
        assert cells, f"{fmt}: ninguna celda con formato de pesos — la prueba no probó nada"
        leaked = [c for c in cells if not isinstance(c[2], str)]
        assert not leaked, f"{fmt}: {leaked[:5]}"
        assert all(value == UNPRICED for _, _, value in cells), fmt

    # The carátula's own four money rows never reach _money_cells: openpyxl
    # only stamps MONEY_FORMAT on a float there, so a leak would be invisible
    # to the sweep above. They are checked by name instead.
    caratula = load_workbook(io.BytesIO(
        build_presupuesto_workbook(report, [], reviews, "Obra", None, fmt="klave")
    ))["Carátula"]
    for label in ("Costo directo", "Precio de venta", "Contingencia",
                  "Total con contingencia"):
        assert _labeled_value(caratula, label, 1, 2) == UNPRICED, label

    for content in (
        build_apus_workbook(report, reviews),
        build_explosion_workbook(report, reviews),
    ):
        workbook = load_workbook(io.BytesIO(content))
        cells = _money_cells(workbook)
        assert cells, workbook.sheetnames
        assert all(value == UNPRICED for _, _, value in cells), workbook.sheetnames


def test_the_same_workbook_does_print_pesos_once_the_unit_is_confirmed():
    """The control for the assertion above: with the unit confirmed the very
    same report resolves to "ok" and every one of those cells carries a
    number again. Without this, a build_presupuesto_workbook that emitted no
    money cells at all — or none of these sheets — would pass."""
    from datetime import UTC, datetime

    report = _doubtful_report()
    reviews = ProjectReviews()
    reviews.verification.units_confirmed_at = datetime.now(UTC)

    workbook = load_workbook(io.BytesIO(
        build_presupuesto_workbook(report, [], reviews, "Obra", None, fmt="klave")
    ))
    printed = _money_cells(workbook)
    sheets = {title for title, _, value in printed if isinstance(value, float)}
    assert {"Carátula", "Presupuesto", "APUs", "Explosión de insumos", "Flujo"} <= sheets
    assert any(
        title == "Presupuesto" and value == report.integration.grand_total
        for title, _, value in printed
    )


def test_a_blocked_workbook_still_carries_its_cantidades():
    """The doctrine withholds the money, not the measurement. A licitación
    sheet with no cantidades would be a different failure, not a fix."""
    report = _doubtful_report()
    quantities = {round(line.quantity, 2) for line in report.boq.lines}
    assert quantities

    ws = load_workbook(io.BytesIO(
        build_presupuesto_workbook(report, [], ProjectReviews(), "Obra", None, fmt="licitacion")
    ))["Catálogo de conceptos"]
    printed = {
        round(c.value, 2) for row in ws.iter_rows() for c in row
        if isinstance(c.value, float)
    }
    assert quantities <= printed


def test_a_lone_heuristic_is_not_enough_to_price():
    units = DrawingUnits(unit="cm", source="text_height_heuristic", confidence=0.5)
    report = generate_cost_report("p", [_wall("w1", 1000.0)], units, CostingConfig(), None, None,
        price_book=LIBRO,
    )
    assert not report.boq.units_reliable and report.boq.direct_cost_total == 0.0


def test_detector_thresholds_follow_the_extent_when_units_are_unknown():
    unknown = DrawingUnits(unit="drawing_units", source="unknown", confidence=0.0)
    generic = DetectorSuiteConfig.preset_for_units(unknown)
    in_cm = DetectorSuiteConfig.preset_for_units(unknown, extent=(0.0, 0.0, 3000.0, 2000.0))
    declared_cm = DetectorSuiteConfig.preset_for_units(
        DrawingUnits(unit="cm", source="dxf_header", confidence=1.0)
    )
    assert in_cm.model_dump() == declared_cm.model_dump()
    assert in_cm.model_dump() != generic.model_dump()
