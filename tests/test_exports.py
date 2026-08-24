"""Deliverable workbooks: structure and evidence-backed generadores."""

import io

from klave_engine.costing.exports import build_presupuesto_workbook
from klave_engine.costing.models import CostingOverrides
from klave_engine.costing.recompute import CostingInputs, build_cost_report
from klave_engine.costing.reviews import ManualAdjustment, ProjectReviews
from klave_engine.detection.results import Detection, DetectionType
from klave_engine.dxf.units import DrawingUnits
from klave_engine.graph.evidence import EvidencePacket
from openpyxl import load_workbook


def _detections() -> list[Detection]:
    return [
        Detection(
            detection_id=f"det_{i}",
            detection_type=DetectionType.column_tag,
            label=f"C{i}",
            bbox=(float(i), 0.0, float(i) + 0.3, 0.3),
            confidence=0.85,
            evidence=EvidencePacket(source="S-101.dxf", method="test"),
            mark=f"C{i}",
            family="columna",
            family_label="Columnas",
            display_label=f"COL-{i:02d}",
        )
        for i in range(1, 4)
    ]


def _report(data_dir):
    inputs = CostingInputs(
        project_id="p",
        detections=_detections(),
        units=DrawingUnits(unit="m", source="dxf_header", confidence=0.9),
        segmentation=None,
        dimensions=None,
    )
    reviews = ProjectReviews(
        adjustments=[
            ManualAdjustment(
                adjustment_id="adj_1",
                concept_code="EST-001",
                quantity_delta=2.5,
                note="castillos del eje 4",
                actor="Diego",
            )
        ]
    )
    return build_cost_report(inputs, CostingOverrides(), reviews=reviews), reviews


def test_klave_workbook_structure_and_generadores(data_dir):
    report, reviews = _report(data_dir)
    content = build_presupuesto_workbook(
        report, _detections(), reviews, "Torre Test", "Cliente SA", fmt="klave"
    )
    workbook = load_workbook(io.BytesIO(content))
    # The four calendarised programs of RLOPSRM art. 45-A-XI travel with the
    # presupuesto: a propuesta that omits them is discarded.
    assert workbook.sheetnames == [
        "Carátula", "Presupuesto", "APUs", "Generadores", "Explosión de insumos", "Programa",
        "Prog. mano de obra", "Prog. maquinaria", "Prog. materiales",
        "Prog. personal técnico", "Flujo",
    ]
    generadores = "\n".join(
        str(cell.value)
        for row in workbook["Generadores"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "COL-" in generadores
    assert "AJUSTE MANUAL +2.50" in generadores
    assert "castillos del eje 4" in generadores
    assert "Diego" in generadores
    caratula = "\n".join(
        str(cell.value)
        for row in workbook["Carátula"].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "SIN VERIFICAR" in caratula


def test_flat_layouts_are_import_friendly(data_dir):
    report, reviews = _report(data_dir)
    for fmt, first_header in (("opus", "Clave"), ("neodata", "Código")):
        content = build_presupuesto_workbook(
            report, _detections(), reviews, "Torre Test", None, fmt=fmt
        )
        ws = load_workbook(io.BytesIO(content))["Presupuesto"]
        assert ws.cell(row=1, column=1).value == first_header
        assert ws.max_row - 1 == len(report.boq.lines)
        assert ws.cell(row=2, column=6).value is not None  # importe populated


def test_apus_workbook_stands_alone(data_dir):
    from klave_engine.costing.exports import build_apus_workbook

    report, _reviews = _report(data_dir)
    workbook = load_workbook(io.BytesIO(build_apus_workbook(report)))
    assert workbook.sheetnames == ["APUs"]
    ws = workbook["APUs"]
    titles = [c.value for c in ws["A"] if isinstance(c.value, str) and " — " in c.value]
    assert len(titles) == len(report.apus) and titles[0].startswith(report.apus[0].concept_code)
    assert any(c.value == "CD unitario" for row in ws.iter_rows() for c in row)


def test_unpriced_lines_say_so_in_every_layout(data_dir):
    from klave_engine.costing.exports import UNPRICED
    from klave_engine.costing.models import BoqLine, QuantityKind

    report, reviews = _report(data_dir)
    report.boq.lines.append(BoqLine(
        concept_code="CIM-010", description="Pilotes (pieza)", unit="PZA", quantity=4.0,
        unit_price=0.0, amount=0.0, unpriced=True, phase="Cimentación", raw_quantity=4.0,
        raw_kind=QuantityKind.COUNT, source_detection_count=4, confidence=0.85,
    ))
    report.boq.totals_by_phase.setdefault("Cimentación", 0.0)
    for fmt in ("klave", "opus", "licitacion"):
        content = build_presupuesto_workbook(
            report, _detections(), reviews, "Torre Test", "Cliente SA", fmt=fmt
        )
        workbook = load_workbook(io.BytesIO(content))
        cells = [
            c.value for ws in workbook.worksheets for row in ws.iter_rows() for c in row
        ]
        assert UNPRICED in cells, fmt
