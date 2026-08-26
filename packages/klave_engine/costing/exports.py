"""Deliverable exports: the documents an estimator actually hands over.

Builds XLSX workbooks from the persisted artifacts:

- ``klave``: carátula, presupuesto formateado, hojas de APU, números
  generadores (the per-concept evidence backup: every detection that produced
  the quantity, plus manual adjustments with note and author), programa y
  flujo. The generadores sheet is the differentiator — it is generated from
  the evidence chain, not typed by hand.
- ``opus`` / ``neodata``: flat, import-wizard-friendly layouts using each
  suite's conventional column order. Both tools import from Excel through a
  column-mapping wizard; these exports keep one row per concept with no
  merged cells so that mapping is trivial.
"""

import io
import math
from collections.abc import Callable
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from klave_engine.costing.descripciones import long_description
from klave_engine.costing.estimaciones import Estimacion, ResumenEstimacion
from klave_engine.costing.explosion import explode
from klave_engine.costing.generadores import calcular as calcular_generador
from klave_engine.costing.letras import pesos_con_letra
from klave_engine.costing.models import BoqLine, Concept, CostReport, MoneyBasis
from klave_engine.costing.presentation import MoneyState, basis_reasons, resolve_money_state
from klave_engine.costing.programas import build_programas
from klave_engine.costing.reviews import ProjectReviews
from klave_engine.costing.schedule import quantity_by_period
from klave_engine.detection.results import Detection

INK = "18181B"
MUTED = "6E6E76"
BORDER_COLOR = "D2D2D7"
SOFT = "F2F2F3"

# A line with quantity and no price: the cell says so instead of a zero
# that would hide in a sum.
UNPRICED = "SIN PRECIO"
# The one cause this sentence actually describes. It used to be emitted for
# every ``blocked`` workbook, and ``blocked`` has three causes — so a run
# whose unit was read as metres at 30 % confidence shipped a catálogo de
# conceptos (LOPSRM art. 45) asserting in writing that its cantidades were
# in unidades de dibujo, directly under a carátula row saying "Unidades del
# plano: m". Every report written before this verdict existed carries no
# basis at all and hit the same sentence.
SIN_UNIDADES = (
    "SIN UNIDADES — la unidad del plano no es confiable: cantidades en unidades de dibujo, "
    "sin precio. Confirma la unidad en Klave antes de usar este archivo."
)
SIN_PRECIOS = "SIN PRECIOS —"
SIN_VEREDICTO = "esta corrida no trae veredicto de unidades"
CANTIDADES_SI = (
    "Las cantidades son las medidas; ningún importe se imprime hasta que confirmes "
    "la unidad en Klave."
)
MONEY_FORMAT = '"$"#,##0.00'
QTY_FORMAT = "#,##0.00"

_thin = Side(style="thin", color=BORDER_COLOR)
_box = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

RAW_KIND_LABELS = {
    "count": "conteo (pza)",
    "length": "longitud (m)",
    "area": "área (m²)",
}

KIND_PROPERTY = {
    "length": ("estimated_span_length", "estimated_length"),
    "area": ("estimated_area",),
}


def _header(ws: Worksheet, row: int, values: list[str]) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.alignment = Alignment(vertical="center")
        cell.border = _box


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _title(ws: Worksheet, row: int, text: str, size: int = 12) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=size, color=INK)


def _muted(ws: Worksheet, row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(size=9, color=MUTED)


def _sentence(text: str) -> str:
    text = text.strip()
    return text if text.endswith((".", "!", "?")) else f"{text}."


def blocked_banner(basis: MoneyBasis | None) -> str:
    """Why this workbook carries no prices, in the run's own words.

    ``resolve_money_state`` returns ``blocked`` for three different reasons
    and only one of them is "the unit is not trustworthy". The reasons were
    already recorded on the run's ``MoneyBasis`` by ``money_basis_from_boq``
    and read back by ``presentation.basis_reasons`` — this is where they were
    always meant to be printed. A basis with no stated reason still gets an
    honest sentence rather than an empty banner.
    """
    if basis is not None and not basis.units_reliable:
        return SIN_UNIDADES
    motivos = " ".join(_sentence(r) for r in basis_reasons(basis) if r.strip())
    return f"{SIN_PRECIOS} {motivos or _sentence(SIN_VEREDICTO)} {CANTIDADES_SI}"


def _banner_row(ws: Worksheet, row: int, report: CostReport) -> None:
    cell = ws.cell(row=row, column=1, value=blocked_banner(report.money_basis))
    cell.font = Font(bold=True, size=9, color="B42318")


def _peso(value: float, blocked: bool) -> float | str:
    """A peso figure, or the honest absence when the verdict withholds money.

    Every derived amount in these workbooks goes through here. Gating only
    the grand total left the presupuesto printing "Precio de venta 29,689.18"
    and "Contingencia 1,484.46" two rows above a withheld TOTAL — the reader
    just adds them. What is gated is money, not the last row.
    """
    return UNPRICED if blocked else value


CroquisProvider = Callable[[BoqLine], list[tuple[str, Path]]]


def build_presupuesto_workbook(
    report: CostReport,
    detections: list[Detection],
    reviews: ProjectReviews,
    project_name: str,
    client: str | None,
    fmt: str = "klave",
    inventory: dict | None = None,
    croquis: CroquisProvider | None = None,
    override_reason: str = "",
) -> bytes:
    # One verdict for the whole workbook, resolved once here and threaded to
    # every sheet that could show a total — not re-derived per sheet from
    # ``report.boq.units_reliable`` alone, which is the bug this module
    # exists to stop repeating.
    state = resolve_money_state(report.money_basis, reviews.verification)
    if fmt == "opus":
        workbook = _flat_workbook(
            report,
            sheet_title="Presupuesto",
            columns=["Clave", "Descripción", "Unidad", "Cantidad", "Precio Unitario", "Importe"],
            money_state=state,
        )
    elif fmt == "neodata":
        workbook = _flat_workbook(
            report,
            sheet_title="Presupuesto",
            columns=["Código", "Concepto", "Unidad", "Cantidad", "P.U.", "Monto"],
            money_state=state,
        )
    elif fmt in ("licitacion", "licitacion_larga"):
        workbook = _licitacion_workbook(
            report, reviews, project_name, client, state,
            long_descriptions=fmt == "licitacion_larga",
        )
    else:
        workbook = _klave_workbook(
            report, detections, reviews, project_name, client, state,
            croquis=croquis,
            override_reason=override_reason,
        )
        if inventory and inventory.get("sheets"):
            _levantamiento(workbook.create_sheet("Levantamiento"), inventory)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ------------------------------------------------------------------ flat ---

IVA_PCT = 16.0


def _licitacion_workbook(
    report: CostReport, reviews: ProjectReviews, project_name: str, client: str | None,
    money_state: MoneyState,
    long_descriptions: bool = False,
) -> Workbook:
    """Catálogo de conceptos for a licitación pública (LOPSRM art. 45 / RLOPSRM
    art. 185): one partida per phase, each concept with its precio unitario
    con número y con letra, importe, subtotal, IVA and total con letra. The
    P.U. is the precio de venta (costo directo × factor de sobrecosto), the
    one a contractor signs — never the bare costo directo."""
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Catálogo de conceptos"
    factor = report.integration.overcost_factor or 1.0
    apus_by_code = {apu.concept_code: apu for apu in report.apus}
    _title(ws, 1, "CATÁLOGO DE CONCEPTOS Y CANTIDADES DE OBRA", size=14)
    _muted(ws, 2, 1, f"Obra: {project_name}")
    _muted(ws, 3, 1, f"Dependencia / cliente: {client or '—'}")
    _muted(
        ws, 4, 1,
        f"Fecha: {datetime.now(UTC):%d/%m/%Y} · Moneda: {report.currency} · "
        f"Precios unitarios con indirectos, financiamiento, utilidad y cargos adicionales "
        f"(factor {factor:.4f} sobre costo directo)",
    )
    verification = reviews.verification
    blocked = money_state == "blocked"
    if blocked:
        _banner_row(ws, 5, report)
    elif not (verification.units_confirmed_at and verification.detections_confirmed_at):
        cell = ws.cell(
            row=5, column=1,
            value="SIN VERIFICAR — cantidades leídas del plano, pendientes de revisión humana",
        )
        cell.font = Font(bold=True, size=9, color="B54708")
    columns = [
        "Partida", "Clave", "Concepto", "Unidad", "Cantidad",
        "P.U. con número", "P.U. con letra", "Importe",
    ]
    _header(ws, 7, columns)
    row = 8
    subtotal = 0.0
    for number, (phase, _phase_total) in enumerate(report.boq.totals_by_phase.items(), 1):
        partida = f"{number:02d}"
        phase_cell = ws.cell(row=row, column=1, value=partida)
        phase_cell.font = Font(bold=True, size=9)
        name_cell = ws.cell(row=row, column=3, value=phase.upper())
        name_cell.font = Font(bold=True, size=9)
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=SOFT)
        row += 1
        partida_total = 0.0
        for index, line in enumerate(
            (ln for ln in report.boq.lines if ln.phase == phase), 1
        ):
            unit_price = round(line.unit_price * factor, 2)
            amount = round(line.quantity * unit_price, 2)
            partida_total += amount
            description = line.description
            if long_descriptions:
                description = long_description(
                    Concept(
                        code=line.concept_code, description=line.description, unit=line.unit,
                        phase=line.phase, production_rate_per_day=1.0,
                    ),
                    apus_by_code.get(line.concept_code),
                )
            sin_precio = blocked or line.unpriced
            values: list[Any] = [
                f"{partida}.{index:03d}", line.taller_clave or line.concept_code,
                description, line.unit, line.quantity,
                UNPRICED if sin_precio else unit_price,
                UNPRICED if sin_precio else pesos_con_letra(unit_price),
                UNPRICED if sin_precio else amount,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = _box
                if col == 5:
                    cell.number_format = QTY_FORMAT
                if col in (6, 8):
                    cell.number_format = MONEY_FORMAT
                if col in (3, 7):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        total_cell = ws.cell(row=row, column=3, value=f"Subtotal partida {partida}")
        total_cell.font = Font(italic=True, size=9, color=MUTED)
        # A withheld SUBTOTAL under three printed partida subtotals is not
        # withheld: the reader adds three numbers.
        amount_cell = ws.cell(row=row, column=8, value=_peso(round(partida_total, 2), blocked))
        amount_cell.number_format = MONEY_FORMAT
        amount_cell.font = Font(italic=True, size=9, color=MUTED)
        subtotal += partida_total
        row += 2
    iva = round(subtotal * IVA_PCT / 100, 2)
    total = round(subtotal + iva, 2)
    # Same rule as every amount above, reaching its last cell: SUBTOTAL/IVA/
    # TOTAL are the closing summary of the sheet this banner already warns
    # about, so they render the absence too rather than the arithmetic.
    for label, value in (
        ("SUBTOTAL", UNPRICED if blocked else round(subtotal, 2)),
        (f"I.V.A. {IVA_PCT:.0f} %", UNPRICED if blocked else iva),
        ("TOTAL", UNPRICED if blocked else total),
    ):
        label_cell = ws.cell(row=row, column=3, value=label)
        label_cell.font = Font(bold=True, size=10)
        value_cell = ws.cell(row=row, column=8, value=value)
        value_cell.number_format = MONEY_FORMAT
        value_cell.font = Font(bold=True, size=10)
        row += 1
    # The spelled-out total is the same number in another alphabet — leaving
    # it would just re-print what the three cells above just withheld.
    letra_text = (
        UNPRICED if blocked else f"Importe total con letra: {pesos_con_letra(total)}"
    )
    letra = ws.cell(row=row, column=3, value=letra_text)
    letra.font = Font(bold=True, size=9)
    row += 2
    _muted(
        ws, row, 1,
        "Las cantidades provienen de la lectura del plano y de las correcciones documentadas "
        "(hoja Generadores del Excel completo); los precios son los del catálogo del taller.",
    )
    widths = {"A": 9, "B": 11, "C": 58, "D": 8, "E": 12, "F": 15, "G": 52, "H": 16}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    return workbook


def _flat_workbook(
    report: CostReport, sheet_title: str, columns: list[str], money_state: MoneyState
) -> Workbook:
    """One row per concept, no merges: made for import wizards."""
    workbook = Workbook()
    ws = workbook.active
    ws.title = sheet_title
    _header(ws, 1, columns)
    row = 2
    blocked = money_state == "blocked"
    if blocked:
        # The first data row carries the warning so an import never reads a
        # doubtful reading as a firm one in silence.
        _banner_row(ws, row, report)
        row += 1
    for line in report.boq.lines:
        # OPUS and Neodata import these columns straight into a presupuesto.
        # A banner they may not read is not a gate; the P.U. and importe
        # columns have to carry the absence themselves.
        sin_precio = blocked or line.unpriced
        values: list[Any] = [
            line.taller_clave or line.concept_code,
            line.description,
            line.unit,
            line.quantity,
            UNPRICED if sin_precio else line.unit_price,
            UNPRICED if sin_precio else line.amount,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _box
            if col == 4:
                cell.number_format = QTY_FORMAT
            if col in (5, 6):
                cell.number_format = MONEY_FORMAT
        row += 1
    _autosize(ws, [14, 64, 10, 14, 16, 18])
    ws.freeze_panes = "A2"
    return workbook


# ----------------------------------------------------------------- klave ---

def _klave_workbook(
    report: CostReport,
    detections: list[Detection],
    reviews: ProjectReviews,
    project_name: str,
    client: str | None,
    money_state: MoneyState,
    croquis: CroquisProvider | None = None,
    override_reason: str = "",
) -> Workbook:
    workbook = Workbook()
    # One verdict, every sheet that prints a peso. The Generadores and
    # Programa sheets take no gate because neither prints one: they carry
    # cantidades, fechas and evidence, and the doctrine keeps those.
    _caratula(workbook.active, report, reviews, project_name, client, money_state, override_reason)
    _presupuesto(workbook.create_sheet("Presupuesto"), report, money_state)
    _apus(workbook.create_sheet("APUs"), report, money_state)
    _generadores(workbook.create_sheet("Generadores"), report, detections, reviews, croquis)
    _explosion(workbook.create_sheet("Explosión de insumos"), report, money_state)
    _programa(workbook.create_sheet("Programa"), report)
    _programas_erogaciones(workbook, report, money_state)
    _flujo(workbook.create_sheet("Flujo"), report, money_state)
    return workbook


def _caratula(
    ws: Worksheet,
    report: CostReport,
    reviews: ProjectReviews,
    project_name: str,
    client: str | None,
    money_state: MoneyState,
    override_reason: str = "",
) -> None:
    ws.title = "Carátula"
    _title(ws, 1, "Presupuesto de obra — Klave", size=16)
    _muted(ws, 2, 1, "Ingeniería de costos con evidencia inspeccionable")

    verification = reviews.verification
    verified = bool(
        verification.units_confirmed_at
        and verification.detections_confirmed_at
        and verification.assumptions_confirmed_at
    )
    blocked = money_state == "blocked"
    rows = [
        ("Proyecto", project_name),
        ("Cliente", client or "—"),
        ("Fecha de emisión", datetime.now(UTC).strftime("%Y-%m-%d")),
        ("Moneda", report.currency),
        ("Unidades del plano", f"{report.drawing_units.unit} "
         f"({report.drawing_units.confidence:.0%} de confianza)"),
        # The banner sits one row under the reading it is about, so it has to
        # agree with it: a row saying "m (30 % de confianza)" above a fixed
        # sentence claiming the drawing is in unidades de dibujo is the screen
        # contradicting itself in the document that gets signed.
        ("Verificación", blocked_banner(report.money_basis) if blocked
         else "Verificado (unidades, detecciones y supuestos)"
         if verified else "SIN VERIFICAR — revisar antes de usar"),
        ("", ""),
        # The four rows this task exists for. Withholding only the last one
        # published its own addends: costo directo + integración = precio de
        # venta, + contingencia = the total, three rows up from where it was
        # supposedly hidden.
        ("Costo directo", _peso(report.boq.direct_cost_total, blocked)),
        ("Precio de venta", _peso(report.integration.sale_price, blocked)),
        ("Contingencia", _peso(report.integration.contingency, blocked)),
        ("Total con contingencia", _peso(report.integration.grand_total, blocked)),
        ("Plazo estimado", f"{report.schedule.total_duration_days} días hábiles"),
    ]
    if override_reason:
        # Exported over a blocking finding: whoever receives this file learns
        # that from the file, not from whoever sent it.
        rows.extend([
            ("", ""),
            ("ENTREGADO CON HALLAZGO BLOQUEANTE", override_reason),
        ])
    for offset, (label, value) in enumerate(rows, start=4):
        label_cell = ws.cell(row=offset, column=1, value=label)
        label_cell.font = Font(bold=True, size=10, color=MUTED)
        value_cell = ws.cell(row=offset, column=2, value=value)
        if isinstance(value, float):
            value_cell.number_format = MONEY_FORMAT
        if label == "Verificación" and not verified:
            value_cell.font = Font(bold=True, color="B54708")
    _muted(
        ws, len(rows) + 5, 1,
        "Precios de referencia salvo indicación contraria; ver la fuente de cada "
        "insumo en el catálogo del taller.",
    )
    _autosize(ws, [26, 46])


def _presupuesto(ws: Worksheet, report: CostReport, money_state: MoneyState) -> None:
    columns = ["Clave", "Concepto", "Unidad", "Cantidad", "P.U. (CD)", "Importe", "Confianza"]
    _header(ws, 1, [*columns, "Por nivel"])
    row = 2
    blocked = money_state == "blocked"
    if blocked:
        _banner_row(ws, row, report)
        row += 1
    for phase, phase_total in report.boq.totals_by_phase.items():
        phase_cell = ws.cell(row=row, column=1, value=phase.upper())
        phase_cell.font = Font(bold=True, size=9, color=MUTED)
        phase_cell.fill = PatternFill("solid", fgColor=SOFT)
        for col in range(2, 9):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=SOFT)
        total_cell = ws.cell(row=row, column=6, value=_peso(phase_total, blocked))
        total_cell.number_format = MONEY_FORMAT
        total_cell.font = Font(bold=True, size=9, color=MUTED)
        total_cell.fill = PatternFill("solid", fgColor=SOFT)
        row += 1
        for line in report.boq.lines:
            if line.phase != phase:
                continue
            sin_precio = blocked or line.unpriced
            values: list[Any] = [
                line.taller_clave or line.concept_code, line.description, line.unit,
                line.quantity,
                UNPRICED if sin_precio else line.unit_price,
                UNPRICED if sin_precio else line.amount,
                f"{line.confidence:.0%}",
                "; ".join(f"{title}: {qty:,.2f}" for title, qty in line.by_view.items()),
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = _box
                if col == 4:
                    cell.number_format = QTY_FORMAT
                if col in (5, 6):
                    cell.number_format = MONEY_FORMAT
            row += 1
    row += 1
    # Every addend, not only the sum: indirectos, financiamiento and utilidad
    # are percentages of a costo directo this sheet would otherwise print in
    # full, so leaving them would reconstruct the withheld TOTAL exactly.
    summary: list[tuple[str, float | str]] = [
        ("Costo directo", _peso(report.boq.direct_cost_total, blocked)),
        *((line.description + f" ({line.percentage}%)", _peso(line.amount, blocked))
          for line in report.integration.lines),
        ("Precio de venta", _peso(report.integration.sale_price, blocked)),
        ("Contingencia", _peso(report.integration.contingency, blocked)),
        ("TOTAL", _peso(report.integration.grand_total, blocked)),
    ]
    for label, amount in summary:
        font = Font(bold=label in ("Costo directo", "Precio de venta", "TOTAL"))
        label_cell = ws.cell(row=row, column=5, value=label)
        label_cell.font = font
        amount_cell = ws.cell(row=row, column=6, value=amount)
        amount_cell.number_format = MONEY_FORMAT
        amount_cell.font = font
        row += 1
    _autosize(ws, [12, 62, 9, 13, 14, 16, 11, 48])
    ws.freeze_panes = "A2"


def _apus(ws: Worksheet, report: CostReport, money_state: MoneyState) -> None:
    row = 1
    blocked = money_state == "blocked"
    if blocked:
        # The matrix is where a P.U. is reconstructed from: consumos × costos.
        # Withholding the presupuesto's importes while this sheet prints every
        # matrix in full hands the reader the arithmetic instead of the answer.
        _banner_row(ws, row, report)
        row += 2
    for apu in report.apus:
        _title(ws, row, f"{apu.concept_code} — {apu.concept_description}", size=11)
        _muted(ws, row + 1, 1, f"Costo directo por {apu.unit}")
        row += 2
        _header(ws, row, ["Recurso", "Descripción", "Unidad", "Cantidad", "Costo", "Importe"])
        row += 1
        if apu.price_source:
            _muted(ws, row, 1, f"P.U. adoptado de {apu.price_source} (sin matriz)")
            row += 1
        for line in apu.lines:
            values: list[Any] = [
                line.resource_code, line.description, line.unit,
                line.quantity, _peso(line.unit_cost, blocked), _peso(line.amount, blocked),
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = _box
                if col == 4:
                    cell.number_format = "#,##0.0000"
                if col in (5, 6):
                    cell.number_format = MONEY_FORMAT
            row += 1
        total_label = ws.cell(row=row, column=5, value="CD unitario")
        total_label.font = Font(bold=True)
        total_cell = ws.cell(row=row, column=6, value=_peso(apu.direct_unit_cost, blocked))
        total_cell.number_format = MONEY_FORMAT
        total_cell.font = Font(bold=True)
        row += 3
    _autosize(ws, [16, 52, 9, 12, 13, 15])


CROQUIS_WIDTH_PX = 520
ROW_PX = 20


def _croquis_rows(ws: Worksheet, row: int, items: list[tuple[str, Path]]) -> int:
    """Each croquis below its concept, captioned with its planta."""
    for title, path in items:
        try:
            image = XlsxImage(str(path))
        except OSError:
            continue
        ratio = CROQUIS_WIDTH_PX / max(image.width, 1)
        image.width, image.height = CROQUIS_WIDTH_PX, int(image.height * ratio)
        _muted(ws, row, 1, f"Croquis · {title}")
        row += 1
        ws.add_image(image, f"A{row}")
        row += int(image.height / ROW_PX) + 2
    return row


def _generadores(
    ws: Worksheet,
    report: CostReport,
    detections: list[Detection],
    reviews: ProjectReviews,
    croquis: CroquisProvider | None = None,
) -> None:
    """The evidence backup sheet: where every quantity comes from — and,
    with a croquis provider, where on the planta it sits."""
    by_id = {d.detection_id: d for d in detections}
    row = 1
    _title(ws, row, "Números generadores", size=13)
    _muted(
        ws, row + 1, 1,
        "Respaldo de cantidades: cada renglón es una detección del plano o un "
        "ajuste manual documentado.",
    )
    row += 3
    for line in report.boq.lines:
        _title(ws, row, f"{line.concept_code} — {line.description}", size=11)
        row += 1
        _muted(
            ws, row, 1,
            f"Base medida: {line.raw_quantity:,.2f} "
            f"{RAW_KIND_LABELS.get(line.raw_kind.value, line.raw_kind.value)} → "
            f"cantidad: {line.quantity:,.2f} {line.unit} · "
            f"confianza {line.confidence:.0%}",
        )
        row += 1
        _header(
            ws, row,
            ["Elemento", "Marca en plano", "Familia", "Hoja", "Medida base", "Confianza"],
        )
        row += 1
        for detection_id in line.source_detections:
            detection = by_id.get(detection_id)
            if detection is None:
                continue
            measure: Any = 1
            for prop in KIND_PROPERTY.get(line.raw_kind.value, ()):
                if detection.properties.get(prop) is not None:
                    measure = round(float(detection.properties[prop]), 3)
                    break
            values: list[Any] = [
                detection.display_label or detection.label,
                detection.mark or "—",
                detection.family_label or detection.detection_type.value,
                detection.evidence.source,
                measure,
                f"{detection.confidence:.0%}",
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = _box
                if col == 5 and isinstance(value, float):
                    cell.number_format = QTY_FORMAT
            row += 1
        # Provenance that is not a detection: mapped levantamiento counts,
        # adopted prices, and the per-planta split.
        for assumption in line.assumptions:
            if assumption.startswith(("Levantamiento:", "P.U. adoptado")):
                cell = ws.cell(row=row, column=1, value=assumption)
                cell.font = Font(italic=True, size=9, color=MUTED)
                row += 1
        if line.by_view:
            split = "; ".join(
                f"{title}: {qty:,.2f} {line.unit}" for title, qty in line.by_view.items()
            )
            cell = ws.cell(row=row, column=1, value=f"Por planta: {split}")
            cell.font = Font(italic=True, size=9, color=MUTED)
            row += 1
        for adjustment in reviews.adjustments:
            if adjustment.concept_code != line.concept_code:
                continue
            if adjustment.quantity_set is not None:
                note = f"CANTIDAD FIJADA EN {adjustment.quantity_set:,.2f} {line.unit}"
                if line.engine_quantity is not None:
                    note += f" (lectura: {line.engine_quantity:,.2f} {line.unit})"
            else:
                note = f"AJUSTE MANUAL {adjustment.quantity_delta:+,.2f} {line.unit}"
            if adjustment.note:
                note += f" — {adjustment.note}"
            if adjustment.actor:
                note += f" ({adjustment.actor})"
            cell = ws.cell(row=row, column=1, value=note)
            cell.font = Font(italic=True, size=9, color="B54708")
            row += 1
        if croquis is not None and line.source_detections:
            row = _croquis_rows(ws, row, croquis(line))
        row += 1
    excluded_summary = [
        (key, review)
        for key, review in reviews.detections.items()
        if review.status == "excluded"
    ]
    if excluded_summary:
        _title(ws, row, "Elementos excluidos por revisión humana", size=11)
        row += 1
        for key, review in excluded_summary:
            note = f"{key}"
            if review.note:
                note += f" — {review.note}"
            if review.actor:
                note += f" ({review.actor})"
            _muted(ws, row, 1, note)
            row += 1
    _autosize(ws, [18, 16, 18, 26, 14, 11])


def _levantamiento(ws: Worksheet, inventory: dict) -> None:
    """Symbols, tags and runs per sheet: the count behind every mapped line
    and the backlog of what is still unmapped."""
    _header(ws, 1, ["Hoja", "Disciplina", "Tipo", "Elemento", "Capa", "Cantidad", "Unidad",
                    "Por planta"])
    row = 2
    unit = inventory.get("unit") or "u. dib."
    for sheet in inventory.get("sheets") or []:
        label = sheet.get("label") or sheet.get("sheet", "")
        discipline = sheet.get("discipline") or ""
        entries: list[list[Any]] = []
        for block in sheet.get("blocks") or []:
            entries.append(["Símbolo", block["block_name"], block.get("layer", ""),
                            block["count"], "PZA", block.get("by_view") or {}])
        for tag in sheet.get("tags") or []:
            entries.append(["Etiqueta", tag["tag"], "", tag["count"], "PZA",
                            tag.get("by_view") or {}])
        for region in sheet.get("areas") or []:
            area = region.get("area_m2")
            entries.append(["Área", region["layer"], region["layer"],
                            area if area is not None else region.get("area_du2", 0.0),
                            "M2" if area is not None else f"{unit}²", region.get("by_view") or {}])
        for run in sheet.get("runs") or []:
            length = run.get("length_m")
            entries.append(["Trazo", run["layer"], run["layer"],
                            length if length is not None else run.get("length_du", 0.0),
                            "M" if length is not None else unit, run.get("by_view") or {}])
        for kind, element, layer, quantity, qty_unit, by_view in entries:
            split = "; ".join(f"{title}: {qty:,.2f}" for title, qty in by_view.items())
            values: list[Any] = [label, discipline, kind, element, layer, quantity, qty_unit, split]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = _box
                if col == 6:
                    cell.number_format = QTY_FORMAT
            row += 1
    row += 1
    for note in inventory.get("notes") or []:
        _muted(ws, row, 1, note)
        row += 1
    _autosize(ws, [34, 14, 10, 34, 24, 12, 8, 48])
    ws.freeze_panes = "A2"


# A drawn bar, not a number to read: the deliverable is a diagrama de barras
# (RLOPSRM art. 45-A-X), so the periods a concept occupies are shaded.
_BAR = PatternFill("solid", fgColor="2B4ACB")
_BAR_CRITICAL = PatternFill("solid", fgColor="B42318")


def _programa(ws: Worksheet, report: CostReport) -> None:
    """Programa de ejecución: the bar chart the convocante asks for, with the
    network behind it (duración, holguras, ruta crítica — art. 224)."""
    schedule = report.schedule
    periods = math.ceil(
        schedule.total_duration_days / max(schedule.workdays_per_month, 1)
    ) if schedule.total_duration_days else 0
    period_headers = [f"Mes {i + 1}" for i in range(periods)]
    _header(ws, 1, [
        "Clave", "Concepto", "Partida", "Cantidad", "Unidad", "Rendimiento/día",
        "Fuente", "Cuadrilla-días", "Inicio", "Término", "Holgura total",
        "Holgura libre", "Ruta crítica", *period_headers,
    ])
    row = 2
    quantities = quantity_by_period(schedule)
    for activity in schedule.activities:
        values: list[Any] = [
            activity.concept_code, activity.description, activity.phase,
            activity.quantity, activity.unit, activity.rendimiento_per_day,
            activity.rendimiento_source, activity.duration_days,
            activity.start_date or activity.start_day,
            activity.end_date or activity.end_day,
            activity.total_float_days, activity.free_float_days,
            "SÍ" if activity.critical else "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _box
            if col == 4:
                cell.number_format = QTY_FORMAT
            if col == 13 and activity.critical:
                cell.font = Font(bold=True, color="B42318")
        # The bar: the quantity this concept executes in each period, shaded.
        per_period = quantities.get(activity.concept_code, [])
        for index in range(periods):
            cell = ws.cell(row=row, column=14 + index)
            cell.border = _box
            value = per_period[index] if index < len(per_period) else 0.0
            if value > 0:
                cell.value = round(value, 2)
                cell.number_format = QTY_FORMAT
                cell.fill = _BAR_CRITICAL if activity.critical else _BAR
                cell.font = Font(color="FFFFFF", size=9)
        row += 1
    row += 1
    ws.cell(row=row, column=2, value="Plazo (días hábiles)").font = Font(bold=True)
    ws.cell(row=row, column=8, value=schedule.total_duration_days).font = Font(bold=True)
    ws.cell(row=row + 1, column=2, value="Plazo contractual (días naturales)").font = Font(
        bold=True
    )
    ws.cell(row=row + 1, column=8, value=schedule.calendar_days).font = Font(bold=True)
    _muted(
        ws, row + 3, 1,
        "Duraciones derivadas del rendimiento de cada matriz (RLOPSRM art. 190). "
        "Red de actividades con holguras y ruta crítica conforme al art. 224. "
        "El plazo contractual se cuenta en días naturales (LOPSRM art. 31 fr. V).",
    )
    _autosize(ws, [12, 46, 14, 12, 9, 14, 10, 13, 12, 12, 12, 12, 12] + [10] * periods)
    ws.freeze_panes = "B2"


def _programas_erogaciones(
    workbook: Workbook, report: CostReport, money_state: MoneyState
) -> None:
    """The four calendarised programs of RLOPSRM art. 45-A-XI, each on its own
    sheet, in its own unit and in pesos.

    The cantidades and the calendar are the programa and they stay; the
    importe column and its totals are money like any other, so a blocked run
    prints the absence there too."""
    programas = build_programas(report)
    blocked = money_state == "blocked"
    period_headers = [
        f"{programas.period_label.capitalize()} {i + 1}" for i in range(programas.periods)
    ]
    for programa in programas.programas:
        title = {
            "mano_de_obra": "Prog. mano de obra",
            "maquinaria": "Prog. maquinaria",
            "materiales": "Prog. materiales",
            "personal_tecnico": "Prog. personal técnico",
        }[programa.rubro]
        ws = workbook.create_sheet(title)
        _title(ws, 1, programa.label, size=12)
        # (d) no lista insumos: lista puestos. El encabezado lo dice, porque
        # "Clave / Insumo" sobre una plantilla de personal se lee mal.
        primeras = (
            ["Tipo", "Puesto", "Unidad", "Personal", "Importe"]
            if programa.rubro == "personal_tecnico"
            else ["Clave", "Insumo", "Unidad", "Cantidad", "Importe"]
        )
        if blocked:
            _banner_row(ws, 2, report)
        _header(ws, 3, [*primeras, *period_headers])
        row = 4
        for entry in programa.rows:
            values: list[Any] = [
                entry.code, entry.description, entry.unit, entry.quantity,
                # Un puesto sin sueldo capturado tiene cantidad y no tiene
                # importe. En cero se leería como que sale gratis.
                "sin sueldo capturado" if entry.sin_importe
                else _peso(entry.amount, blocked),
                *entry.by_period,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = _box
                if col == 4 or col > 5:
                    cell.number_format = QTY_FORMAT
                if col == 5 and not entry.sin_importe:
                    cell.number_format = MONEY_FORMAT
            row += 1
        if programa.rows:
            ws.cell(row=row, column=2, value="TOTAL (importe)").font = Font(bold=True)
            total_cell = ws.cell(row=row, column=5, value=_peso(programa.total, blocked))
            total_cell.font = Font(bold=True)
            total_cell.number_format = MONEY_FORMAT
            # These are pesos per period, not cantidades — the row above them
            # is what carries the calendarised quantity.
            for index, value in enumerate(programa.total_by_period):
                cell = ws.cell(row=row, column=6 + index, value=_peso(value, blocked))
                cell.font = Font(bold=True)
                cell.number_format = MONEY_FORMAT
            row += 1
        for note in programa.notes:
            row += 1
            _muted(ws, row, 1, note)
        _autosize(ws, [14, 46, 10, 14, 16] + [14] * programas.periods)


def _flujo(ws: Worksheet, report: CostReport, money_state: MoneyState) -> None:
    _header(ws, 1, ["Periodo", "Avance %", "Gasto directo", "Estimación",
                    "Amort. anticipo", "Retención", "Flujo neto", "Fact. acumulada"])
    row = 2
    blocked = money_state == "blocked"
    if blocked:
        # Six of this sheet's eight columns are pesos, and the last one
        # accumulates to the withheld total by construction. The calendar and
        # the avance curve stay: those are the programa, not the money.
        _banner_row(ws, row, report)
        row += 1
    for period in report.financial.periods:
        values: list[Any] = [
            period.label, period.progress_pct,
            _peso(period.direct_spend, blocked), _peso(period.billing, blocked),
            _peso(period.advance_amortization, blocked), _peso(period.retention, blocked),
            _peso(period.net_cashflow, blocked), _peso(period.accumulated_billing, blocked),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _box
            if col == 2:
                cell.number_format = "0.0"
            if col >= 3:
                cell.number_format = MONEY_FORMAT
        row += 1
    _autosize(ws, [12, 10, 16, 16, 16, 14, 16, 17])
    ws.freeze_panes = "A2"


def _explosion(ws: Worksheet, report: CostReport, money_state: MoneyState) -> None:
    """Every resource the presupuesto consumes: quantity, cost, by partida."""
    explosion = explode(report)
    blocked = money_state == "blocked"
    _title(ws, 1, "Explosión de insumos", size=13)
    _muted(
        ws, 2, 1,
        "APU × cantidad, sumado por insumo. Lo que hay que comprar, contratar y programar; "
        "los conceptos con P.U. adoptado sin matriz no se explotan.",
    )
    if blocked:
        # "Total explotado" is the costo directo under another name, and the
        # per-insumo importes rebuild it. The cantidades and the per-partida
        # split are cantidades — they stay, and they are the useful half of
        # this sheet for buying and contracting.
        _banner_row(ws, 3, report)
    phases = list(report.boq.totals_by_phase.keys())
    _header(
        ws, 4,
        ["Insumo", "Descripción", "Tipo", "Unidad", "Cantidad", "Costo unitario", "Importe",
         *phases],
    )
    row = 5
    type_label = {"material": "Material", "mano_de_obra": "Mano de obra", "equipo": "Equipo"}
    for r in explosion.resources:
        values: list[Any] = [
            r.code, r.description, type_label.get(r.resource_type, r.resource_type), r.unit,
            # by_phase is a cantidad per partida, not an importe: it stays.
            r.quantity, _peso(r.unit_cost, blocked), _peso(r.amount, blocked),
            *[r.by_phase.get(phase, 0.0) or None for phase in phases],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _box
            if col == 5 or col > 7:
                cell.number_format = QTY_FORMAT
            if col in (6, 7):
                cell.number_format = MONEY_FORMAT
        row += 1
    row += 1
    for label, value in (
        ("Materiales", explosion.by_type.get("material", 0.0)),
        ("Mano de obra", explosion.by_type.get("mano_de_obra", 0.0)),
        ("Equipo y herramienta", explosion.by_type.get("equipo", 0.0)),
        ("Total explotado", explosion.total),
    ):
        ws.cell(row=row, column=2, value=label).font = Font(bold=True, size=9)
        cell = ws.cell(row=row, column=7, value=_peso(value, blocked))
        cell.number_format = MONEY_FORMAT
        cell.font = Font(bold=True, size=9)
        row += 1
    for note in explosion.notes:
        _muted(ws, row, 1, note)
        row += 1
    for letter, width in {"A": 18, "B": 48, "C": 13, "D": 8, "E": 13, "F": 14, "G": 15}.items():
        ws.column_dimensions[letter].width = width


def build_apus_workbook(report: CostReport, reviews: ProjectReviews) -> bytes:
    """Every APU as it prints in the presupuesto workbook, on its own.

    Takes ``reviews`` for the same reason ``build_presupuesto_workbook`` does:
    the verdict is a join of what the engine read (on the report) and what a
    person signed off (in the reviews), and a download that skips the second
    half is a download with no gate."""
    workbook = Workbook()
    _apus(workbook.active, report, resolve_money_state(report.money_basis, reviews.verification))
    workbook.active.title = "APUs"
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_explosion_workbook(report: CostReport, reviews: ProjectReviews) -> bytes:
    workbook = Workbook()
    _explosion(
        workbook.active, report, resolve_money_state(report.money_basis, reviews.verification)
    )
    workbook.active.title = "Explosión de insumos"
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_cotizacion_workbook(ages: list) -> bytes:
    """Solicitud de cotización: one row per insumo with its current price and
    vigencia, and the supplier's columns empty. Re-imports as cotización."""
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Solicitud de cotización"
    _title(ws, 1, "SOLICITUD DE COTIZACIÓN DE INSUMOS", size=14)
    _muted(
        ws, 2, 1,
        f"Emitida {datetime.now(UTC):%d/%m/%Y}. Favor de llenar 'Precio cotizado' (MXN, sin IVA), "
        "'Proveedor' y 'Vigencia'; el archivo se importa tal cual en Klave.",
    )
    _header(
        ws, 4,
        ["Clave", "Descripción", "Unidad", "Precio actual", "Vigencia actual", "Estado",
         "Precio cotizado", "Proveedor", "Vigencia", "Observaciones"],
    )
    row = 5
    for age in ages:
        values: list[Any] = [
            age.code, age.description, age.unit, age.unit_cost, age.vigencia or "—",
            age.status, None, None, None, None,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _box
            if col == 4:
                cell.number_format = MONEY_FORMAT
            if col == 7:
                cell.number_format = MONEY_FORMAT
                cell.fill = PatternFill("solid", fgColor="FFF7E6")
        row += 1
    for letter, width in {"A": 18, "B": 56, "C": 8, "D": 14, "E": 14, "F": 10, "G": 16,
                          "H": 24, "I": 12, "J": 30}.items():
        ws.column_dimensions[letter].width = width
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_estimacion_workbook(
    estimacion: Estimacion, resumen: ResumenEstimacion, obra: str = ""
) -> bytes:
    """Una estimación como se entrega: carátula, conceptos y generadores.

    El orden no es decorativo. La carátula va primero porque es lo que se firma;
    los conceptos después, con lo anterior y lo acumulado a la vista para que se
    pueda seguir la cuenta de un periodo al siguiente; y los generadores al
    final, que es el respaldo que se revisa cuando un número no convence.

    Los tres van en la misma hoja y no en tres archivos, porque una estimación
    que llega sin su generador se regresa, y separarlos es la forma más fácil de
    que uno de los dos se quede en el escritorio.
    """
    workbook = Workbook()
    ws = workbook.active
    ws.title = f"Estimación {estimacion.numero}"

    _title(ws, 1, f"ESTIMACIÓN {estimacion.numero}", size=14)
    _muted(ws, 2, 1, f"{obra} · periodo {resumen.periodo}" if obra else resumen.periodo)

    row = 4
    _title(ws, row, "Carátula", size=11)
    row += 1
    for etiqueta, valor, signo in [
        ("Ejecutado en el periodo", resumen.importe, 1),
        (f"Amortización del anticipo ({estimacion.anticipo_pct:g} %)", resumen.amortizacion, -1),
        (
            f"Retención del fondo de garantía ({estimacion.retencion_pct:g} %)",
            resumen.retencion,
            -1,
        ),
        ("Deductivas", resumen.deductivas, -1),
        ("LÍQUIDO A PAGAR", resumen.liquido, 0),
    ]:
        ws.cell(row=row, column=1, value=etiqueta).font = Font(
            bold=signo == 0, size=10, color=INK
        )
        cell = ws.cell(row=row, column=3, value=round(valor * (signo if signo else 1), 2))
        cell.number_format = MONEY_FORMAT
        cell.font = Font(bold=signo == 0, size=10)
        row += 1

    # Cada deductiva con su nombre: un descuento sin razón es el que se reclama.
    if estimacion.deductivas:
        row += 1
        _title(ws, row, "Deductivas aplicadas", size=11)
        row += 1
        _header(ws, row, ["Concepto", "Importe", "Razón"])
        row += 1
        for d in estimacion.deductivas:
            ws.cell(row=row, column=1, value=d.concepto).border = _box
            imp = ws.cell(row=row, column=2, value=round(d.importe, 2))
            imp.number_format = MONEY_FORMAT
            imp.border = _box
            ws.cell(row=row, column=3, value=d.razon or "—").border = _box
            row += 1

    row += 1
    _title(ws, row, "Conceptos estimados", size=11)
    row += 1
    _header(ws, row, [
        "Clave", "Concepto", "Unidad", "P.U.", "Contratado",
        "Anterior", "Este periodo", "Acumulado", "% avance", "Importe",
    ])
    row += 1
    for r in estimacion.renglones:
        valores: list[Any] = [
            r.clave, r.description, r.unit, r.unit_price, r.quantity_contract,
            r.quantity_previous, r.quantity_period, r.quantity_accumulated,
            round(r.pct_avance / 100.0, 4), r.amount_period,
        ]
        for col, value in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _box
            if col in (4, 10):
                cell.number_format = MONEY_FORMAT
            elif col == 9:
                cell.number_format = "0.0%"
        if r.excede_contrato:
            # Rojo en la celda del acumulado: es la que delata que hace falta
            # convenio, y quien revisa busca ahí.
            ws.cell(row=row, column=8).font = Font(size=10, color="B42318", bold=True)
        row += 1

    row += 2
    _title(ws, row, "Números generadores", size=11)
    row += 1
    _muted(
        ws, row, 1,
        "El respaldo de lo medido en obra. Un concepto sin generador se cobra sin "
        "sustento y la estimación se regresa (RLOPSRM art. 132).",
    )
    row += 2

    for r in estimacion.renglones:
        if r.quantity_period <= 0:
            continue
        _title(ws, row, f"{r.clave} — {r.description}", size=10)
        row += 1
        if not r.generador:
            _muted(
                ws, row, 1,
                f"Sin generador capturado. Se cobran {r.quantity_period:,.4g} {r.unit}.",
            )
            row += 2
            continue

        g = calcular_generador(r.generador, r.unit, r.quantity_period)
        _header(ws, row, ["Ubicación", "Veces", "Operación", "Medida", "Nota"])
        row += 1
        for calc in g.lineas:
            ws.cell(row=row, column=1, value=calc.linea.ubicacion or "—").border = _box
            ws.cell(row=row, column=2, value=calc.linea.veces).border = _box
            # La fórmula viaja escrita: quien revisa rehace la cuenta, no la cree.
            ws.cell(row=row, column=3, value=calc.formula or "—").border = _box
            medida = ws.cell(
                row=row, column=4,
                value=calc.medida if calc.medida is not None else f"falta {', '.join(calc.falta)}",
            )
            medida.border = _box
            if calc.medida is None:
                medida.font = Font(size=10, color="B42318")
            ws.cell(row=row, column=5, value=calc.linea.nota or "").border = _box
            row += 1

        total = ws.cell(row=row, column=3, value="Suma del generador")
        total.font = Font(bold=True, size=10)
        suma = ws.cell(row=row, column=4, value=g.total)
        suma.font = Font(bold=True, size=10)
        row += 1
        if not g.cuadra:
            _muted(ws, row, 1, g.avisos[-1] if g.avisos else "")
            ws.cell(row=row, column=1).font = Font(size=9, color="B42318")
            row += 1
        row += 1

    if resumen.avisos:
        row += 1
        _title(ws, row, "Avisos", size=11)
        row += 1
        for aviso in resumen.avisos:
            _muted(ws, row, 1, aviso)
            row += 1

    _autosize(ws, [26, 46, 10, 14, 14, 14, 14, 14, 11, 16])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
