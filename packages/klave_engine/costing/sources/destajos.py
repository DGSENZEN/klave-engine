"""Matrices y cuadrillas de un catálogo de destajos, por bloques.

Un catálogo de destajo mexicano no se escribe como tabla plana: se escribe
como una matriz por bloque, que es como se lee en papel.

    CLAVE: N1-ALB-MUR-01          UNIDAD: m2      CANTIDAD: 1
    MURO DE CONTENCIÓN DE 12 CM …
    CLAVE INSUMO | DESCRIPCIÓN | UNIDAD | CANTIDAD | COSTO UNITARIO | IMPORTE
    1F1A   | CUADRILLA NO 6 (1 FIERRERO + …) | JOR   | 0.022589 | 1524.63
    %MO1   | HERRAMIENTA MENOR               | (%)MO | 0.03     |  375.27
    COSTO DIRECTO TOTAL

Las hojas de cuadrillas traen la misma forma, un nivel más abajo: la
cuadrilla es el bloque y sus integrantes son los renglones, cada uno con su
categoría y su jornal. De ahí sale el costo de la mano de obra que las
matrices usan, sin que nadie lo invente.

**Los materiales aparecen en cero, y así hay que dejarlos.** No es un hueco
del archivo: un destajo paga la mano de obra y la obra pone el material, así
que el catálogo lista el concreto como insumo y le pone cero porque el
destajista no lo cobra. Rellenar ese cero con un precio de otro lado
convertiría un catálogo honesto en uno inventado, que es justo lo que este
módulo existe para evitar.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

from klave_engine.costing.sources.custom import CustomCatalogError
from klave_engine.costing.sources.matrices import ConceptRow, InsumoRow, MatricesParse

# El encabezado de un bloque: «CLAVE: X … UNIDAD: u … CANTIDAD: n». Hay hojas
# que lo escriben «MATRIZ: X» y otras que ponen la unidad en su propia celda.
_CLAVE = re.compile(r"^\s*(?:CLAVE|MATRIZ)\s*:\s*(\S[^|]*?)\s*$", re.IGNORECASE)
_UNIDAD = re.compile(r"^\s*UNIDAD\s*:\s*(\S+)", re.IGNORECASE)
_ENCABEZADO_INSUMOS = re.compile(r"CLAVE\s+(INSUMO|INTEGRANTE)|^\s*CLAVE\s*$", re.IGNORECASE)
_FIN = re.compile(r"COSTO\s+DIRECTO|TOTAL\s+DE\s+|^\s*SUBTOTAL", re.IGNORECASE)
# Renglones que son rótulos de sección dentro de la matriz, no insumos.
_SECCION = re.compile(
    r"^\s*(MANO\s+DE\s+OBRA|MATERIALES?|HERRAMIENTA|EQUIPO|AUXILIARES?|B[AÁ]SICOS?)\s*$",
    re.IGNORECASE,
)
_PORCENTAJE = re.compile(r"%\s*\)?\s*MO|\(\s*%\s*\)", re.IGNORECASE)
# La hoja de cuadrillas define mano de obra compuesta, no obra vendible: una
# cuadrilla es un insumo con su jornal, y su clave —1A, MOCU-003— se repite
# igual en las tres zonas con precios distintos. Importarlas como conceptos
# hacía que la última zona pisara a la anterior y que el catálogo del taller
# se llenara de renglones que nadie puede presupuestar.
_HOJA_CUADRILLAS = re.compile(r"CUADRILLA", re.IGNORECASE)
# Una unidad de obra suelta en su celda, cuando la hoja no la rotula.
_UNIDAD_SUELTA = re.compile(
    r"m2|m3|ml|m|pza|pieza|jor|kg|ton|lote|sal|salida|hr|hora|juego|lt|l",
    re.IGNORECASE,
)
# Los catálogos que rotulan la descripción con «Concepto:» delante.
_PREFIJO_DESCRIPCION = re.compile(r"^\s*concepto\s*:\s*", re.IGNORECASE)


def _texto(valor: object) -> str:
    return "" if valor is None else str(valor).strip()


def _numero(valor: object) -> float | None:
    if valor is None:
        return None
    if isinstance(valor, int | float):
        return float(valor)
    limpio = re.sub(r"[^\d.,-]", "", str(valor)).replace(",", "")
    try:
        return float(limpio)
    except ValueError:
        return None


def _tipo_de_insumo(clave: str, unidad: str, descripcion: str) -> tuple[str, bool]:
    """Material, mano de obra o equipo, por lo que dice la clave y la unidad."""
    if _PORCENTAJE.search(unidad):
        return "equipo", True
    material = f"{clave} {descripcion}".upper()
    if re.search(r"CUADRILLA|^MO\d|^MOCU|^CUA-|PEON|OFICIAL|AYUDANTE|CABO", material):
        return "mano_de_obra", False
    if re.search(r"EQUIPO|M[AÁ]QUINA|REVOLVEDORA|VIBRADOR|COMPACTADOR|ANDAMIO", material):
        return "equipo", False
    if unidad.upper().startswith("JOR") or unidad.upper() in {"HORA", "HR", "H"}:
        return "mano_de_obra", False
    return "material", False


@dataclass
class _Bloque:
    clave: str
    unidad: str
    descripcion: str = ""
    renglones: list[tuple[str, str, str, float, float]] = field(default_factory=list)


def _bloques(filas: list[list[object]]) -> list[_Bloque]:
    """Cada matriz de la hoja, con sus renglones."""
    salida: list[_Bloque] = []
    actual: _Bloque | None = None
    esperando_descripcion = False
    for fila in filas:
        celdas = [_texto(v) for v in fila]
        unido = " ".join(c for c in celdas if c)
        if not unido:
            continue
        cabecera = next((_CLAVE.match(c) for c in celdas if _CLAVE.match(c)), None)
        if cabecera is not None:
            if actual is not None and actual.renglones:
                salida.append(actual)
            unidad = next((m.group(1) for c in celdas if (m := _UNIDAD.match(c))), "")
            if not unidad:
                # Hay hojas que no rotulan la unidad: la ponen en su propia
                # columna después de la clave —«MATRIZ: X | | m2 | 1»—, así
                # que se toma la primera celda que parezca una unidad de obra.
                unidad = next(
                    (c for c in celdas[1:] if _UNIDAD_SUELTA.fullmatch(c)), ""
                )
            actual = _Bloque(clave=cabecera.group(1).strip(), unidad=unidad)
            esperando_descripcion = True
            continue
        if actual is None:
            continue
        if _FIN.search(unido):
            if actual.renglones:
                salida.append(actual)
            actual = None
            continue
        if _ENCABEZADO_INSUMOS.search(unido):
            esperando_descripcion = False
            continue
        if _SECCION.match(unido):
            continue
        if esperando_descripcion and not _numero(celdas[3] if len(celdas) > 3 else None):
            actual.descripcion = _PREFIJO_DESCRIPCION.sub("", unido).strip()
            esperando_descripcion = False
            continue
        # Un renglón de insumo: clave, descripción, unidad, cantidad, costo.
        if len(celdas) < 5 or not celdas[0]:
            continue
        cantidad, costo = _numero(celdas[3]), _numero(celdas[4])
        if cantidad is None or cantidad <= 0 or costo is None:
            continue
        actual.renglones.append((celdas[0], celdas[1], celdas[2], cantidad, costo))
    if actual is not None and actual.renglones:
        salida.append(actual)
    return salida


def parse_destajos_workbook(raw: bytes, filename: str = "") -> MatricesParse:
    """Conceptos con su matriz real, de un catálogo de destajos por bloques.

    Recorre todas las hojas: las de cuadrillas definen el costo de la mano de
    obra y las de matrices lo usan, así que las dos importan y en ese orden
    no hace falta — cada bloque trae el costo de su renglón escrito."""
    if not (filename.lower().endswith((".xlsx", ".xlsm")) or raw[:2] == b"PK"):
        raise CustomCatalogError("El catálogo de destajos debe ser un XLSX.")
    from openpyxl import load_workbook

    try:
        libro = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise CustomCatalogError("No se pudo leer el XLSX.") from exc

    conceptos: list[ConceptRow] = []
    insumos: dict[str, InsumoRow] = {}
    problemas: list[str] = []
    vistos: set[str] = set()
    for hoja in libro.worksheets:
        filas = [list(r) for r in hoja.iter_rows(values_only=True)]
        es_cuadrilla = bool(_HOJA_CUADRILLAS.search(hoja.title or ""))
        for bloque in _bloques(filas):
            if es_cuadrilla:
                # La cuadrilla entra como insumo, con el jornal que suman sus
                # integrantes. Es lo que las matrices consumen.
                jornal = round(
                    sum(cant * costo for _c, _d, _u, cant, costo in bloque.renglones), 2
                )
                if jornal <= 0:
                    continue
                previo = insumos.get(bloque.clave)
                if previo is None or previo.unit_cost <= 0:
                    insumos[bloque.clave] = InsumoRow(
                        code=bloque.clave,
                        description=bloque.descripcion or f"Cuadrilla {bloque.clave}",
                        unit=(bloque.unidad or "JOR").upper(),
                        unit_cost=jornal, resource_type="mano_de_obra",
                    )
                continue
            if bloque.clave in vistos:
                continue
            vistos.add(bloque.clave)
            # Un insumo repetido dentro de una matriz es el mismo insumo: se
            # suman sus cantidades. Insertarlo dos veces no es «más detalle»,
            # es la misma cuadrilla contada dos veces.
            acumulado: dict[str, float] = {}
            orden: list[str] = []
            for clave, descripcion, unidad, cantidad, costo in bloque.renglones:
                tipo, porcentaje = _tipo_de_insumo(clave, unidad, descripcion)
                previo = insumos.get(clave)
                if previo is None or (costo > 0 and previo.unit_cost <= 0):
                    insumos[clave] = InsumoRow(
                        code=clave, description=descripcion or clave, unit=unidad or "JOR",
                        unit_cost=costo, resource_type=tipo, is_labor_percentage=porcentaje,
                    )
                if clave not in acumulado:
                    orden.append(clave)
                acumulado[clave] = round(acumulado.get(clave, 0.0) + cantidad, 8)
            componentes = [(c, acumulado[c]) for c in orden]
            if not componentes:
                problemas.append(f"{bloque.clave}: sin renglones utilizables.")
                continue
            conceptos.append(
                ConceptRow(
                    code=bloque.clave,
                    description=bloque.descripcion or bloque.clave,
                    unit=(bloque.unidad or "").upper() or "PZA",
                    phase="Destajos",
                    production_rate_per_day=None,
                    components=componentes,
                )
            )
    libro.close()
    if not conceptos:
        raise CustomCatalogError(
            "No se encontró ninguna matriz con la forma «CLAVE: … / insumos / COSTO DIRECTO»."
        )
    return MatricesParse(concepts=conceptos, insumos=insumos, problems=problemas)
