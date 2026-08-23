"""Conceptos con sus matrices (APUs) from an OPUS / Neodata Excel export.

Both tools export a "catálogo de conceptos con insumos" as a flat sheet: a
concept row (clave, descripción, unidad — and optionally its rendimiento
and partida) followed by the rows of its matriz (insumo clave,
descripción, unidad, cantidad por unidad de concepto, costo unitario).
Neither native format is open; the Excel export is what every taller
already has. This parser reads that layout and the explicit one Klave
documents (a "Tipo" column saying CONCEPTO / INSUMO), tolerant of header
names in either tool's vocabulary.

A concept needs a clave, descripción, unidad and at least one insumo with
a positive quantity; an insumo needs a clave, unidad and a positive cost
(a cost of zero is "por cotización", not a price). Anything else is
reported, never silently imported.
"""

from __future__ import annotations

import csv
import io
import re
from collections.abc import Iterable
from dataclasses import dataclass, field

from klave_engine.costing.sources.custom import (
    CustomCatalogError,
    _matches,
    _norm,
    _price,
)

TIPO_HEADERS = ("tipo", "tipo de renglon", "tipo de registro", "clase", "nivel")
CLAVE_HEADERS = ("clave", "codigo", "código", "code", "key", "id", "clave del concepto")
DESCRIPTION_HEADERS = ("descripcion", "descripción", "concepto", "description", "nombre")
UNIT_HEADERS = ("unidad", "unit", "u", "um", "u.m.")
QUANTITY_HEADERS = ("cantidad", "cant", "cant.", "quantity", "rendimiento de insumo",
                    "cantidad por unidad", "cantidad/unidad")
COST_HEADERS = ("costo", "costo unitario", "precio", "precio unitario", "p.u.", "pu",
                "costo directo", "unit cost", "importe unitario")
RATE_HEADERS = ("rendimiento", "rend.", "rend", "rendimiento diario", "unidades/dia",
                "unidades/día", "avance diario")
PHASE_HEADERS = ("partida", "fase", "capitulo", "capítulo", "grupo", "etapa")

CONCEPT_WORDS = ("concepto", "c", "matriz", "basico", "básico", "compuesto", "analisis",
                 "análisis")
INSUMO_WORDS = ("insumo", "i", "recurso", "material", "mano de obra", "mo", "equipo",
                "herramienta", "eq", "mat", "auxiliar")
LABOR_WORDS = ("mano de obra", "mo", "cuadrilla", "oficial", "peon", "peón", "ayudante")
EQUIP_WORDS = ("equipo", "eq", "herramienta", "maquinaria", "costo horario")


@dataclass
class InsumoRow:
    code: str
    description: str
    unit: str
    unit_cost: float
    resource_type: str  # material | mano_de_obra | equipo
    is_labor_percentage: bool = False


@dataclass
class ConceptRow:
    code: str
    description: str
    unit: str
    phase: str
    production_rate_per_day: float | None
    components: list[tuple[str, float]] = field(default_factory=list)


@dataclass
class MatricesParse:
    concepts: list[ConceptRow]
    insumos: dict[str, InsumoRow]
    problems: list[str]


def _find_header(rows: list[list[object]]) -> tuple[int, dict[str, int]] | None:
    for index, row in enumerate(rows[:80]):
        cells = ["" if v is None else str(v) for v in row]
        columns: dict[str, int] = {}
        for position, cell in enumerate(cells):
            if not cell.strip():
                continue
            for key, candidates in (
                ("tipo", TIPO_HEADERS), ("clave", CLAVE_HEADERS),
                ("description", DESCRIPTION_HEADERS), ("unit", UNIT_HEADERS),
                ("quantity", QUANTITY_HEADERS), ("cost", COST_HEADERS),
                ("rate", RATE_HEADERS), ("phase", PHASE_HEADERS),
            ):
                if key not in columns and _matches(cell, candidates):
                    columns[key] = position
                    break
        if {"clave", "description", "unit", "quantity"} <= set(columns):
            return index, columns
    return None


def _resource_type(tipo: str, code: str, unit: str, description: str) -> tuple[str, bool]:
    text = _norm(f"{tipo} {description}")
    normalized_code = code.upper()
    if unit.strip() in ("%", "%MO", "% MO"):
        return "equipo", True
    if normalized_code.startswith(("MO", "MOB")) or any(w in text.split() for w in ("mo",)):
        return "mano_de_obra", False
    if any(w in text for w in LABOR_WORDS):
        return "mano_de_obra", False
    if normalized_code.startswith(("EQ", "HER", "MAQ")) or any(w in text for w in EQUIP_WORDS):
        return "equipo", False
    return "material", False


def _is_concept(tipo: str, quantity: float | None, cost: float | None) -> bool:
    normalized = _norm(tipo)
    if normalized:
        if normalized in CONCEPT_WORDS or normalized.startswith("concep"):
            return True
        if normalized in INSUMO_WORDS or normalized.startswith("insu"):
            return False
    # Without a Tipo column a concept row has no quantity (its cost is the
    # sum of its matriz), an insumo row has one.
    return quantity is None


def _quantity(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    text = str(value).strip().replace(" ", "")
    if re.fullmatch(r"\d+,\d{1,4}", text):
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return number if number > 0 else None


def parse_matrices_table(table: list[list[object]]) -> MatricesParse:
    found = _find_header(table)
    if found is None:
        raise CustomCatalogError(
            "No se encontró un encabezado con clave, descripción, unidad y cantidad "
            "(exporta desde OPUS/Neodata el catálogo de conceptos con sus insumos)."
        )
    start, columns = found

    def cell(row: list[object], key: str) -> object:
        position = columns.get(key)
        return row[position] if position is not None and position < len(row) else None

    concepts: list[ConceptRow] = []
    insumos: dict[str, InsumoRow] = {}
    problems: list[str] = []
    current: ConceptRow | None = None
    phase = "Sin partida"
    for number, row in enumerate(table[start + 1:], start + 2):
        clave = str(cell(row, "clave") or "").strip()
        description = " ".join(str(cell(row, "description") or "").split())
        unit = str(cell(row, "unit") or "").strip().upper().replace("M²", "M2").replace("M³", "M3")
        tipo = str(cell(row, "tipo") or "").strip()
        quantity = _quantity(cell(row, "quantity"))
        cost = _price(cell(row, "cost"))
        rate = _quantity(cell(row, "rate"))
        phase_value = str(cell(row, "phase") or "").strip()
        if phase_value:
            phase = phase_value[:60]
        if not clave and not description:
            continue
        if not unit and not quantity and description and not clave:
            phase = description[:60]  # a partida heading row
            continue
        if _is_concept(tipo, quantity, cost):
            if not clave or not unit:
                problems.append(f"Fila {number}: concepto sin clave o unidad ({description[:40]}).")
                current = None
                continue
            current = ConceptRow(
                code=clave.upper(), description=description or clave, unit=unit,
                phase=phase, production_rate_per_day=rate,
            )
            concepts.append(current)
            continue
        # An insumo row.
        if current is None:
            problems.append(
                f"Fila {number}: insumo {clave or description[:30]} sin concepto arriba."
            )
            continue
        if not clave:
            problems.append(f"Fila {number}: insumo sin clave bajo {current.code}.")
            continue
        if quantity is None:
            problems.append(f"Fila {number}: insumo {clave} sin cantidad bajo {current.code}.")
            continue
        code = clave.upper()
        resource_type, is_pct = _resource_type(tipo, code, unit, description)
        if code not in insumos:
            if cost is None and not is_pct:
                problems.append(
                    f"Fila {number}: insumo {code} sin costo (por cotización); la matriz de "
                    f"{current.code} queda sin ese recurso."
                )
                continue
            insumos[code] = InsumoRow(
                code=code, description=description or code, unit=unit or "PZA",
                unit_cost=cost if cost is not None else quantity, resource_type=resource_type,
                is_labor_percentage=is_pct,
            )
        current.components.append((code, quantity))
    concepts = [c for c in concepts if c.components] or concepts
    empty = [c.code for c in concepts if not c.components]
    for code in empty:
        problems.append(f"Concepto {code} sin insumos: no se importa (sin matriz no hay precio).")
    concepts = [c for c in concepts if c.components]
    if not concepts:
        raise CustomCatalogError("Ningún concepto trae insumos con cantidad; nada que importar.")
    return MatricesParse(concepts=concepts, insumos=insumos, problems=problems)


def parse_matrices_workbook(raw: bytes, filename: str = "") -> MatricesParse:
    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm")) or raw[:2] == b"PK":
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise CustomCatalogError("No se pudo leer el XLSX.") from exc
        last_error: CustomCatalogError | None = None
        for sheet in workbook.worksheets:
            table = [list(r) for r in sheet.iter_rows(values_only=True)]
            try:
                return parse_matrices_table(table)
            except CustomCatalogError as exc:
                last_error = exc
        raise last_error or CustomCatalogError("El libro no tiene hojas.")
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = raw.decode("cp1252")
        except UnicodeDecodeError as exc:
            raise CustomCatalogError("El archivo debe ser CSV UTF-8/CP1252 o XLSX.") from exc
    delimiter = ";" if text.count(";") > text.count(",") else ","
    csv_table: list[list[object]] = [
        list(r) for r in csv.reader(io.StringIO(text), delimiter=delimiter)
    ]
    return parse_matrices_table(csv_table)


def iter_components(parse: MatricesParse) -> Iterable[tuple[str, list[tuple[str, float]]]]:
    for concept in parse.concepts:
        yield concept.code, list(concept.components)
