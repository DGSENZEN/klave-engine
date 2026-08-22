"""A taller's own catálogo de conceptos (XLSX/CSV) as a reference source.

Offices keep their price book as a workbook: one row per concept with a
clave, a description, a unit and a precio unitario, grouped under partida
headings (PRELIMINARES, ALBAÑILERÍA…). The layout varies — column order,
header wording, a title block above the header — so the parser finds the
header row by its words and reads whatever follows. Rows without a price
are headings or "por cotización" and become the group of the rows below.
"""

import csv
import io
import re
import unicodedata
from collections.abc import Iterable

CLAVE_HEADERS = ("clave", "codigo", "código", "code", "key", "clave destajo inteligente")
DESCRIPTION_HEADERS = (
    "descripcion", "descripción", "description", "concepto", "descripcion tecnica del concepto",
    "descripción técnica del concepto",
)
UNIT_HEADERS = ("unidad", "unit", "u.m.", "um", "und")
PRICE_HEADERS = (
    "precio unitario", "precio", "p.u.", "pu", "p. u.", "precio_unitario", "unit_cost",
    "costo unitario", "costo", "p.u. mano de obra", "importe unitario",
)


class CustomCatalogError(ValueError):
    pass


def _norm(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.lower().split())


def _matches(header: str, candidates: Iterable[str]) -> bool:
    normalized = _norm(header)
    return any(normalized == _norm(c) or normalized.startswith(_norm(c) + " ") for c in candidates)


def _price(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    text = str(value).strip().replace("$", "").replace(" ", "")
    if not text:
        return None
    # "1,250.50" is thousands+decimal; "412,50" (no dot) is a decimal comma.
    if "," in text and "." not in text and re.fullmatch(r"\d+,\d{1,2}", text):
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return number if number > 0 else None


def _find_header(rows: list[list[object]]) -> tuple[int, dict[str, int]] | None:
    for index, row in enumerate(rows[:60]):
        cells = ["" if v is None else str(v) for v in row]
        columns: dict[str, int] = {}
        for position, cell in enumerate(cells):
            if not cell.strip():
                continue
            if "clave" not in columns and _matches(cell, CLAVE_HEADERS):
                columns["clave"] = position
            elif "description" not in columns and _matches(cell, DESCRIPTION_HEADERS):
                columns["description"] = position
            elif "unit" not in columns and _matches(cell, UNIT_HEADERS):
                columns["unit"] = position
            elif "price" not in columns and _matches(cell, PRICE_HEADERS):
                columns["price"] = position
        if {"clave", "description", "unit", "price"} <= set(columns):
            return index, columns
    return None


def _rows_from_table(table: list[list[object]]) -> list[dict]:
    found = _find_header(table)
    if found is None:
        return []
    start, columns = found
    rows: list[dict] = []
    group_clave = ""
    group_description = ""

    def cell(row: list[object], key: str) -> object:
        position = columns[key]
        return row[position] if position < len(row) else None

    for row in table[start + 1:]:
        clave = str(cell(row, "clave") or "").strip()
        description = " ".join(str(cell(row, "description") or "").split())
        unit = str(cell(row, "unit") or "").strip()
        price = _price(cell(row, "price"))
        if not clave and not description:
            continue
        if price is None or not unit:
            # A partida heading (PRELIMINARES) or a concept "por cotización".
            heading = description or clave
            if heading and not unit:
                group_clave, group_description = (clave if description else ""), heading[:600]
            continue
        if not clave:
            continue
        rows.append(
            {
                "clave": clave,
                "description": description or clave,
                "unit": unit.upper().replace("M²", "M2").replace("M³", "M3"),
                "price": price,
                "group_clave": group_clave,
                "group_description": group_description,
            }
        )
    return rows


def parse_concept_workbook(raw: bytes, filename: str = "") -> list[dict]:
    """Concept rows from an XLSX (every sheet with a recognizable header) or
    a CSV. Raises CustomCatalogError when no header with clave, descripción,
    unidad and precio exists."""
    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm")) or raw[:2] == b"PK":
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise CustomCatalogError("No se pudo leer el XLSX.") from exc
        rows: list[dict] = []
        for sheet in workbook.worksheets:
            table = [list(r) for r in sheet.iter_rows(values_only=True)]
            rows.extend(_rows_from_table(table))
    else:
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = raw.decode("cp1252")
            except UnicodeDecodeError as exc:
                raise CustomCatalogError("El archivo debe ser CSV UTF-8/CP1252 o XLSX.") from exc
        delimiter = ";" if text.count(";") > text.count(",") else ","
        table = [list(r) for r in csv.reader(io.StringIO(text), delimiter=delimiter)]
        rows = _rows_from_table(table)
    if not rows:
        raise CustomCatalogError(
            "No se encontró un encabezado con clave, descripción, unidad y precio unitario."
        )
    return rows


def source_key_for(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", _norm(name)).strip("-")[:48] or "catalogo"
    return f"propio-{slug}"
