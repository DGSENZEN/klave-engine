"""A past presupuesto as a file: clave, descripción, unidad, cantidad and —
when present — precio unitario. The taller's history, readable from any
XLSX/CSV they already have (OPUS/Neodata exports, their own generadores)."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass

from klave_engine.costing.sources.custom import (
    CustomCatalogError,
    _matches,
    _norm,
    _price,
)

CLAVE_HEADERS = ("clave", "codigo", "código", "code", "key", "id")
DESCRIPTION_HEADERS = ("descripcion", "descripción", "concepto", "description", "nombre")
UNIT_HEADERS = ("unidad", "unit", "u", "um", "u.m.")
QUANTITY_HEADERS = ("cantidad", "cant", "cant.", "quantity", "volumen")
PRICE_HEADERS = ("precio", "precio unitario", "p.u.", "pu", "p. u.", "costo", "costo unitario",
                 "unit price")


@dataclass
class PresupuestoRow:
    clave: str
    description: str
    unit: str
    quantity: float
    price: float | None
    group: str = ""


def _quantity(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    text = str(value).strip().replace(" ", "")
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return number if number > 0 else None


def _find_header(rows: list[list[object]]) -> tuple[int, dict[str, int]] | None:
    for index, row in enumerate(rows[:80]):
        cells = ["" if v is None else str(v) for v in row]
        columns: dict[str, int] = {}
        for position, cell in enumerate(cells):
            if not cell.strip():
                continue
            for key, candidates in (
                ("clave", CLAVE_HEADERS), ("description", DESCRIPTION_HEADERS),
                ("unit", UNIT_HEADERS), ("quantity", QUANTITY_HEADERS),
                ("price", PRICE_HEADERS),
            ):
                if key not in columns and _matches(cell, candidates):
                    columns[key] = position
                    break
        if {"clave", "unit", "quantity"} <= set(columns):
            return index, columns
    return None


def parse_presupuesto_table(table: list[list[object]]) -> list[PresupuestoRow]:
    found = _find_header(table)
    if found is None:
        raise CustomCatalogError(
            "No se encontró un encabezado con clave, unidad y cantidad (y de preferencia "
            "descripción y precio unitario)."
        )
    start, columns = found

    def cell(row: list[object], key: str) -> object:
        position = columns.get(key)
        return row[position] if position is not None and position < len(row) else None

    rows: list[PresupuestoRow] = []
    group = ""
    for row in table[start + 1:]:
        clave = str(cell(row, "clave") or "").strip().upper()
        description = " ".join(str(cell(row, "description") or "").split())
        unit = str(cell(row, "unit") or "").strip().upper().replace("M²", "M2").replace("M³", "M3")
        quantity = _quantity(cell(row, "quantity"))
        price = _price(cell(row, "price"))
        if not clave and not description:
            continue
        if quantity is None or not unit:
            heading = description or clave
            if heading and not unit:
                group = heading[:60]
            continue
        if not clave:
            continue
        rows.append(
            PresupuestoRow(
                clave=clave, description=description or clave, unit=unit,
                quantity=quantity, price=price, group=group,
            )
        )
    if not rows:
        raise CustomCatalogError("El presupuesto no trae renglones con clave, unidad y cantidad.")
    return rows


def parse_presupuesto_file(raw: bytes, filename: str = "") -> list[PresupuestoRow]:
    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm")) or raw[:2] == b"PK":
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise CustomCatalogError("No se pudo leer el XLSX.") from exc
        last: CustomCatalogError | None = None
        for sheet in workbook.worksheets:
            table = [list(r) for r in sheet.iter_rows(values_only=True)]
            try:
                return parse_presupuesto_table(table)
            except CustomCatalogError as exc:
                last = exc
        raise last or CustomCatalogError("El libro no tiene hojas.")
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = raw.decode("cp1252")
        except UnicodeDecodeError as exc:
            raise CustomCatalogError("El archivo debe ser CSV UTF-8/CP1252 o XLSX.") from exc
    delimiter = ";" if text.count(";") > text.count(",") else ","
    csv_table: list[list[object]] = [
        list(r) for r in csv.reader(io.StringIO(text), delimiter=delimiter)
    ]
    return parse_presupuesto_table(csv_table)


def normalized_clave(clave: str) -> str:
    return _norm(clave).upper().replace(" ", "")
