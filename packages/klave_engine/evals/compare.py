"""Compare the engine's presupuesto against a human one — the calibration a
firm runs on its first projects to decide how far to trust the numbers.

    uv run python -m klave_engine.evals.compare COST_REPORT.json HUMANO.xlsx
    uv run python -m klave_engine.evals.compare PROJECT_ROOT HUMANO.xlsx

The human file is any XLSX/CSV with clave, descripción, unidad and cantidad
columns (OPUS/Neodata exports, a taller's own presupuesto). Concepts are
matched by clave; lines the human has and the engine lacks — and the
reverse — are listed, never hidden. Output: a Markdown table with the
quantity delta per concept and a summary (coverage, median |Δ|, lines
within ±10 %), written to stdout and optionally to a file.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import statistics
import sys
from dataclasses import dataclass
from pathlib import Path

from klave_engine.costing.matching import Candidate, rank, unit_key
from klave_engine.costing.sources.custom import _find_header, _matches, _norm

QUANTITY_HEADERS = ("cantidad", "cant", "cant.", "quantity", "volumen")
AMOUNT_HEADERS = ("importe", "monto", "total", "amount")


@dataclass
class HumanLine:
    clave: str
    description: str
    unit: str
    quantity: float
    amount: float | None = None


@dataclass
class Comparison:
    clave: str
    description: str
    unit: str
    engine: float | None
    human: float | None
    engine_amount: float | None = None
    human_amount: float | None = None
    # "clave" (exact or taller alias), "descripción NN %" (matcher), or "".
    matched_by: str = ""
    unit_mismatch: str = ""  # the engine's unit when it disagrees with the human's

    @property
    def delta_pct(self) -> float | None:
        if self.engine is None or self.human in (None, 0) or self.unit_mismatch:
            return None
        assert self.human is not None
        return (self.engine - self.human) / self.human * 100.0


def _table_from_file(path: Path) -> list[list[object]]:
    raw = path.read_bytes()
    if path.suffix.lower() in (".xlsx", ".xlsm") or raw[:2] == b"PK":
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        table: list[list[object]] = []
        for sheet in workbook.worksheets:
            table.extend(list(r) for r in sheet.iter_rows(values_only=True))
        return table
    text = raw.decode("utf-8-sig", errors="replace")
    delimiter = ";" if text.count(";") > text.count(",") else ","
    return [list(r) for r in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _number(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace("$", "")
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def read_human_presupuesto(path: Path) -> list[HumanLine]:
    """Lines with clave, descripción, unidad and cantidad; the price column is
    optional (a cantidades-only generador is enough to compare quantities)."""
    table = _table_from_file(path)
    found = _find_header(table)
    columns: dict[str, int] = {}
    start = 0
    if found is not None:
        start, columns = found
    # The custom header finder needs a price column; a cantidades-only file
    # does not have one, so find the quantity column ourselves.
    for index, row in enumerate(table[:60]):
        cells = ["" if v is None else str(v) for v in row]
        if any(_matches(c, QUANTITY_HEADERS) for c in cells if c.strip()):
            if found is None:
                for position, header_cell in enumerate(cells):
                    lowered = _norm(header_cell)
                    if not lowered:
                        continue
                    if "clave" not in columns and lowered in ("clave", "codigo", "code"):
                        columns["clave"] = position
                    elif "description" not in columns and lowered.startswith(
                        ("descrip", "concepto")
                    ):
                        columns["description"] = position
                    elif "unit" not in columns and lowered in ("unidad", "unit", "u", "um"):
                        columns["unit"] = position
                start = index
            columns["quantity"] = next(
                p for p, c in enumerate(cells) if c.strip() and _matches(c, QUANTITY_HEADERS)
            )
            for position, header_cell in enumerate(cells):
                if header_cell.strip() and _matches(header_cell, AMOUNT_HEADERS):
                    columns["amount"] = position
                    break
            break
    if not {"clave", "quantity"} <= set(columns):
        raise SystemExit(
            "El presupuesto humano necesita columnas de clave y cantidad (y de preferencia "
            "descripción y unidad)."
        )

    def cell(row: list[object], key: str) -> object:
        position = columns.get(key)
        return row[position] if position is not None and position < len(row) else None

    lines: list[HumanLine] = []
    for row in table[start + 1:]:
        clave = str(cell(row, "clave") or "").strip().upper()
        quantity = _number(cell(row, "quantity"))
        if not clave or quantity is None:
            continue
        lines.append(
            HumanLine(
                clave=clave,
                description=" ".join(str(cell(row, "description") or "").split())[:80],
                unit=str(cell(row, "unit") or "").strip().upper(),
                quantity=quantity,
                amount=_number(cell(row, "amount")),
            )
        )
    return lines


def load_engine_lines(source: Path) -> list[dict]:
    """From a cost_report.json, or from a project root's active run."""
    if source.is_dir():
        pointer = source / "processed" / "active_run.json"
        candidates = []
        if pointer.exists():
            run = json.loads(pointer.read_text())["artifact_dir"]
            candidates.append(source / "processed" / run / "cost_report.json")
        candidates.append(source / "processed" / "cost_report_override.json")
        candidates.append(source / "processed" / "cost_report.json")
        source = next((c for c in candidates if c.exists()), candidates[0])
    report = json.loads(source.read_text())
    return list(report["boq"]["lines"])


def _comparison(h: HumanLine, e: dict | None, matched_by: str = "") -> Comparison:
    unit_mismatch = ""
    if e is not None and h.unit and unit_key(h.unit) != unit_key(str(e["unit"])):
        unit_mismatch = str(e["unit"])
    return Comparison(
        clave=h.clave,
        description=h.description or (e["description"][:80] if e else ""),
        unit=h.unit or (e["unit"] if e else ""),
        engine=float(e["quantity"]) if e else None,
        human=h.quantity,
        engine_amount=float(e.get("amount") or 0.0) or None if e else None,
        human_amount=h.amount,
        matched_by=matched_by,
        unit_mismatch=unit_mismatch,
    )


def compare(engine_lines: list[dict], human: list[HumanLine]) -> list[Comparison]:
    """Three passes, honestly labeled: the exact clave (theirs or the taller
    alias the engine already prints), then the description matcher for what
    is left, then the orphans on both sides."""
    by_clave: dict[str, dict] = {}
    for line in engine_lines:
        by_clave.setdefault(str(line["concept_code"]).upper(), line)
        taller = str(line.get("taller_clave") or "").upper()
        if taller:
            by_clave.setdefault(taller, line)

    matched_engine: set[int] = set()
    out: list[Comparison] = []
    unmatched_humans: list[HumanLine] = []
    for h in human:
        e = by_clave.get(h.clave)
        if e is not None:
            matched_engine.add(id(e))
            out.append(_comparison(h, e, matched_by="clave"))
        else:
            unmatched_humans.append(h)

    # Second pass: descriptions. A human line without a clave the engine
    # knows may still be the same concept written the taller's way.
    remaining = [line for line in engine_lines if id(line) not in matched_engine]
    candidates = [
        Candidate(
            kind="concept", key=str(index), clave=str(line["concept_code"]),
            description=str(line["description"]), unit=str(line["unit"]), price=None,
        )
        for index, line in enumerate(remaining)
    ]
    for h in unmatched_humans:
        matches = rank(h.description, h.unit, candidates, limit=1) if h.description else []
        best = matches[0] if matches and matches[0].score >= 0.6 else None
        if best is not None:
            line = remaining[int(best.candidate.key)]
            if id(line) not in matched_engine:
                matched_engine.add(id(line))
                out.append(_comparison(h, line, matched_by=f"descripción {best.score:.0%}"))
                continue
        out.append(_comparison(h, None))

    for line in engine_lines:
        if id(line) not in matched_engine:
            out.append(
                Comparison(
                    clave=str(line["concept_code"]), description=line["description"][:80],
                    unit=str(line["unit"]), engine=float(line["quantity"]), human=None,
                    engine_amount=float(line.get("amount") or 0.0) or None,
                )
            )
    return out


def render_markdown(rows: list[Comparison]) -> str:
    matched = [r for r in rows if r.delta_pct is not None]
    deltas = [abs(r.delta_pct or 0.0) for r in matched]
    within = sum(1 for d in deltas if d <= 10.0)
    lines = [
        "| Clave | Concepto | Unidad | Motor | Humano | Δ % | Emparejado por |",
        "|---|---|---|---:|---:|---:|---|",
    ]
    for r in sorted(rows, key=lambda r: (r.delta_pct is None, -abs(r.delta_pct or 0.0))):
        engine = f"{r.engine:,.2f}" if r.engine is not None else "—"
        human = f"{r.human:,.2f}" if r.human is not None else "—"
        if r.unit_mismatch:
            delta = f"unidad distinta ({r.unit_mismatch} vs {r.unit})"
        elif r.delta_pct is not None:
            delta = f"{r.delta_pct:+.1f}"
        else:
            delta = "solo humano" if r.engine is None else "solo motor"
        lines.append(
            f"| {r.clave} | {r.description} | {r.unit} | {engine} | {human} | {delta} "
            f"| {r.matched_by or '—'} |"
        )
    only_human = sum(1 for r in rows if r.engine is None)
    only_engine = sum(1 for r in rows if r.human is None)
    mismatched_units = sum(1 for r in rows if r.unit_mismatch)
    summary = [
        "",
        f"Conceptos comparados: {len(matched)} · solo en el humano: {only_human} · "
        f"solo en el motor: {only_engine}"
        + (f" · unidades en desacuerdo: {mismatched_units}" if mismatched_units else ""),
    ]
    if deltas:
        summary.append(
            f"Mediana |Δ|: {statistics.median(deltas):.1f} % · dentro de ±10 %: "
            f"{within}/{len(deltas)} · peor: {max(deltas):.1f} %"
        )
    engine_total = sum(r.engine_amount or 0.0 for r in rows)
    human_total = sum(r.human_amount or 0.0 for r in rows)
    if engine_total and human_total:
        drift = (engine_total - human_total) / human_total * 100.0
        summary.append(
            f"Importe (líneas con monto): motor ${engine_total:,.2f} vs humano "
            f"${human_total:,.2f} ({drift:+.1f} %)"
        )
    return "\n".join(lines + summary)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("engine", type=Path, help="cost_report.json o raíz del proyecto")
    parser.add_argument("human", type=Path, help="presupuesto humano (XLSX/CSV)")
    parser.add_argument("--out", type=Path, default=None, help="archivo Markdown de salida")
    args = parser.parse_args(argv)
    rows = compare(load_engine_lines(args.engine), read_human_presupuesto(args.human))
    text = render_markdown(rows)
    if args.out:
        args.out.write_text(text + "\n", encoding="utf-8")
    sys.stdout.write(text + "\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
